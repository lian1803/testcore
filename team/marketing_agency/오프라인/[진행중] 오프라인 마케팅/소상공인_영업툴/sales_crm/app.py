#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
오프라인 마케팅팀 영업 CRM
- PDF 자동 감시 → 업체 추가
- 개별 채팅 + 거절 응답 AI 생성
- 대화 기록 저장
- naver-diagnosis message_generator 연동
"""

import sys
import io
import warnings

# 경고 숨기기
warnings.filterwarnings("ignore")

# UTF-8 인코딩 설정 (PyInstaller 호환)
try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
except (AttributeError, TypeError):
    pass  # PyInstaller EXE에서 실패해도 계속 진행

from flask import Flask, render_template, request, jsonify, send_file
import anthropic
import google.generativeai as genai
import os
import sqlite3
import json
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
import threading
import time
import webbrowser
import tempfile
import re
import subprocess

load_dotenv()
# 폴백: 메인 .env에서도 로드 (exe에서 .env 못 찾을 경우)
_REPO_ROOT = Path(__file__).parent.parent.parent.parent.parent
load_dotenv(str(_REPO_ROOT / "company" / ".env"))

# 환경변수 설정 (없으면 기본값)
os.environ.setdefault("SALES_REGION", "양주")
os.environ.setdefault("SENDER_NAME", "리안")
os.environ.setdefault("SENDER_PHONE", "010-XXXX-XXXX")

app = Flask(__name__, template_folder='templates', static_folder='static')
anthropic_client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

# 설정
_SALES_TOOL_DIR = Path(__file__).parent.parent  # 소상공인_영업툴/
PDF_FOLDER = _SALES_TOOL_DIR / "naver-diagnosis"
NAVER_DIAGNOSIS_DB = _SALES_TOOL_DIR / "naver-diagnosis" / "diagnosis.db"
EXCEL_DB_FOLDER = _SALES_TOOL_DIR.parent / "소상공인 010수집 최종본"
DB_PATH = "sales_crm.db"

# naver-diagnosis 모듈 직접 로드 (services/__init__.py 거치지 않음 → playwright 의존성 회피)
NAVER_DIAG_PATH = _SALES_TOOL_DIR / "naver-diagnosis"
sys.path.insert(0, str(NAVER_DIAG_PATH))
try:
    import importlib.util as _ilu

    _mg_spec = _ilu.spec_from_file_location("message_generator", str(NAVER_DIAG_PATH / "services" / "message_generator.py"))
    _mg_mod = _ilu.module_from_spec(_mg_spec)
    _mg_spec.loader.exec_module(_mg_mod)
    msg_gen_first = _mg_mod.generate_first_message
    msg_gen_fourth = _mg_mod.generate_fourth_messages
    MESSAGE_GENERATOR_AVAILABLE = True

    _hp_spec = _ilu.spec_from_file_location("html_pdf_generator", str(NAVER_DIAG_PATH / "services" / "html_pdf_generator.py"))
    _hp_mod = _ilu.module_from_spec(_hp_spec)
    _hp_spec.loader.exec_module(_hp_mod)
    HtmlPdfGenerator = _hp_mod.HtmlPdfGenerator
    HTML_PDF_AVAILABLE = True
except Exception as e:
    print(f"⚠️ 모듈 임포트 실패: {e}")
    MESSAGE_GENERATOR_AVAILABLE = False
    HTML_PDF_AVAILABLE = False

# ──────────────────────────────────────────────────
# naver-diagnosis DB 조회 함수
# ──────────────────────────────────────────────────

def get_diagnosis_data(business_name: str) -> dict:
    """
    naver-diagnosis 진단 데이터 조회
    업체명으로 검색해서 가장 최근 진단 데이터를 message_generator 형식으로 변환

    매칭 전략:
    1. 정확 매칭: WHERE business_name = ? (완전한 데이터 우선 — competitor_avg_review > 0)
    2. 공백 제거 매칭: REPLACE(business_name, ' ', '') = REPLACE(?, ' ', '')
    3. 부분 매칭: WHERE business_name LIKE ?
    4. 지역명 제거 매칭: 업체명에서 지역명 제거 후 재매칭

    Args:
        business_name: 업체명

    Returns:
        message_generator.generate_first_message()가 요구하는 dict 형식
        또는 기본 데이터 (DB에 없는 경우)
    """
    try:
        if not NAVER_DIAGNOSIS_DB.exists():
            print(f"⚠️ naver-diagnosis DB 파일 없음: {NAVER_DIAGNOSIS_DB}")
            return None

        conn = sqlite3.connect(str(NAVER_DIAGNOSIS_DB))
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        row = None

        # 1단계: 정확 매칭 (완전한 데이터 우선 — 경쟁사 데이터 있는 것 먼저)
        c.execute("""
            SELECT * FROM diagnosis_history
            WHERE business_name = ? AND competitor_avg_review > 0
            ORDER BY created_at DESC
            LIMIT 1
        """, (business_name,))
        row = c.fetchone()

        # 1-1단계: 정확 매칭이지만 경쟁사 데이터 없으면 그냥 최신
        if not row:
            c.execute("""
                SELECT * FROM diagnosis_history
                WHERE business_name = ?
                ORDER BY created_at DESC
                LIMIT 1
            """, (business_name,))
            row = c.fetchone()

        # 2단계: 공백 제거 매칭 (1단계 실패 시) — 역시 완전한 데이터 우선
        if not row:
            c.execute("""
                SELECT * FROM diagnosis_history
                WHERE REPLACE(business_name, ' ', '') = REPLACE(?, ' ', '') AND competitor_avg_review > 0
                ORDER BY created_at DESC
                LIMIT 1
            """, (business_name,))
            row = c.fetchone()

        # 2-1단계: 공백 제거 매칭이지만 경쟁사 데이터 없으면 그냥 최신
        if not row:
            c.execute("""
                SELECT * FROM diagnosis_history
                WHERE REPLACE(business_name, ' ', '') = REPLACE(?, ' ', '')
                ORDER BY created_at DESC
                LIMIT 1
            """, (business_name,))
            row = c.fetchone()

        # 3단계: 부분 매칭 (1,2단계 실패 시) — 완전한 데이터 우선
        if not row:
            c.execute("""
                SELECT * FROM diagnosis_history
                WHERE business_name LIKE ? AND competitor_avg_review > 0
                ORDER BY created_at DESC
                LIMIT 1
            """, (f"%{business_name}%",))
            row = c.fetchone()

        # 3-1단계: 부분 매칭이지만 경쟁사 데이터 없으면 그냥 최신
        if not row:
            c.execute("""
                SELECT * FROM diagnosis_history
                WHERE business_name LIKE ?
                ORDER BY created_at DESC
                LIMIT 1
            """, (f"%{business_name}%",))
            row = c.fetchone()

        # 4단계: 지역명 제거 매칭 (1,2,3단계 실패 시) — 역시 완전한 데이터 우선
        if not row:
            regions = ['양주', '포천', '의정부', '동두천', '연천', '가평', '남양주']
            for region in regions:
                cleaned = business_name.replace(region, '').strip()
                if cleaned and cleaned != business_name:
                    c.execute("""
                        SELECT * FROM diagnosis_history
                        WHERE business_name LIKE ? AND competitor_avg_review > 0
                        ORDER BY created_at DESC
                        LIMIT 1
                    """, (f'%{cleaned}%',))
                    row = c.fetchone()
                    if row:
                        break

        # 4-1단계: 지역명 제거 매칭이지만 경쟁사 데이터 없으면 그냥 최신
        if not row:
            regions = ['양주', '포천', '의정부', '동두천', '연천', '가평', '남양주']
            for region in regions:
                cleaned = business_name.replace(region, '').strip()
                if cleaned and cleaned != business_name:
                    c.execute("""
                        SELECT * FROM diagnosis_history
                        WHERE business_name LIKE ?
                        ORDER BY created_at DESC
                        LIMIT 1
                    """, (f'%{cleaned}%',))
                    row = c.fetchone()
                    if row:
                        break

        conn.close()

        if not row:
            print(f"⚠️ DB에 {business_name} 진단 데이터 없음. 기본 데이터로 생성")
            return None

        # dict로 변환
        data = dict(row)

        # JSON 필드 파싱
        if isinstance(data.get('keywords'), str):
            try:
                data['keywords'] = json.loads(data['keywords'])
            except:
                data['keywords'] = []

        if isinstance(data.get('improvement_points'), str):
            try:
                data['improvement_points'] = json.loads(data['improvement_points'])
            except:
                data['improvement_points'] = []

        return data

    except Exception as e:
        print(f"⚠️ naver-diagnosis DB 조회 실패: {e}")
        return None


def infer_category_from_name(name: str) -> str:
    """업체명에서 카테고리 추정"""
    keywords = {
        "미용실": ["헤어", "미용", "매직", "파마", "스타일"],
        "네일": ["네일", "손톱", "페디"],
        "식당": ["음식", "밥", "국수", "카레", "숯불"],
        "카페": ["카페", "커피", "아메리카노", "라떼"],
        "피부관리": ["피부", "관리", "에스테틱", "스킨", "여드름"],
        "학원": ["학원", "영어", "수학", "과외", "튜터"],
    }

    name_lower = name.lower()
    for category, words in keywords.items():
        if any(word in name_lower for word in words):
            return category

    return "기타"

def get_default_diagnosis_data(business_name: str, phone: str) -> dict:
    """기본 진단 데이터 생성 (DB 조회 실패 시)"""
    return {
        "business_name": business_name,
        "category": infer_category_from_name(business_name),
        "review_count": 0,
        "photo_count": 0,
        "keyword_score": 30,
        "review_score": 20,
        "blog_score": 10,
        "info_score": 25,
        "photo_score": 15,
        "convenience_score": 25,
        "engagement_score": 10,
        "total_score": 18,
        "grade": "D",
        "naver_place_rank": 15,
        "competitor_avg_review": 0,
        "competitor_avg_photo": 0,
        "estimated_lost_customers": 0,
        "has_kakao": False,
        "has_instagram": False,
        "has_hours": False,
        "has_menu": False,
        "has_price": False,
        "has_intro": False,
        "has_directions": False,
        "has_owner_reply": False,
        "news_last_days": 90,
        "keywords": [],
    }


# ──────────────────────────────────────────────────
# DB 초기화
# ──────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # 업체 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS businesses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        phone TEXT UNIQUE,
        name TEXT,
        location TEXT,
        pdf_path TEXT,
        place_url TEXT,
        status TEXT DEFAULT '1차_발송_대기',
        naver_diagnosis_id INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    # 대화 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS conversations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        business_id INTEGER,
        speaker TEXT,
        message TEXT,
        message_order INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(business_id) REFERENCES businesses(id)
    )''')

    # 멘트 시퀀스 테이블 (1차~4차 메시지 저장)
    c.execute('''CREATE TABLE IF NOT EXISTS message_sequences (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        business_id INTEGER,
        sequence_num INTEGER,
        message TEXT,
        sent_at TIMESTAMP,
        customer_response TEXT,
        customer_responded_at TIMESTAMP,
        FOREIGN KEY(business_id) REFERENCES businesses(id)
    )''')

    conn.commit()
    conn.close()

init_db()


def migrate_db():
    """앱 시작 시 DB 컬럼 자동 동기화 — 코드에 컬럼 추가해도 DB 에러 안 남"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("PRAGMA table_info(businesses)")
    existing = {row[1] for row in c.fetchall()}

    # 없는 컬럼 자동 추가
    migrations = [
        ("place_url",            "TEXT"),
        ("naver_diagnosis_id",   "INTEGER"),
        ("memo",                 "TEXT"),
    ]
    for col, col_type in migrations:
        if col not in existing:
            c.execute(f"ALTER TABLE businesses ADD COLUMN {col} {col_type}")
            print(f"✅ DB 마이그레이션: businesses.{col} 추가")

    conn.commit()
    conn.close()

migrate_db()


def sync_from_naver_diagnosis():
    """
    naver-diagnosis DB에서 진단 완료된 업체를 CRM DB로 자동 동기화.
    앱 시작 시 + PDF 감시 루프에서 주기적으로 호출.
    """
    try:
        if not NAVER_DIAGNOSIS_DB.exists():
            return 0

        naver_conn = sqlite3.connect(str(NAVER_DIAGNOSIS_DB))
        naver_conn.row_factory = sqlite3.Row
        nc = naver_conn.cursor()

        nc.execute("""SELECT id, business_name, address, category, place_url,
                      photo_count, review_count, total_score, grade,
                      ppt_filename, priority_tag
                      FROM diagnosis_history ORDER BY created_at DESC""")
        rows = nc.fetchall()
        naver_conn.close()

        if not rows:
            return 0

        crm_conn = sqlite3.connect(DB_PATH)
        cc = crm_conn.cursor()
        added = 0

        for row in rows:
            name = row["business_name"]
            if not name:
                continue

            # 이미 CRM에 있는지 확인 (업체명 기준)
            cc.execute("SELECT id FROM businesses WHERE name = ?", (name,))
            if cc.fetchone():
                continue

            # Excel에서 진짜 전화번호 조회
            excel_phone, excel_addr = get_phone_from_excel(name)
            phone_val = excel_phone if excel_phone else f"진단-{row['id']}"
            addr_val = excel_addr or row["address"] or ""

            # PDF 경로: 빈 문자열로 저장 (나중에 PDF 다운로드 시 동적 생성)
            cc.execute("""INSERT OR IGNORE INTO businesses (phone, name, location, pdf_path, status)
                         VALUES (?, ?, ?, ?, ?)""",
                       (phone_val, name, addr_val, "", "1차_발송_대기"))

            business_id = cc.lastrowid
            added += 1

        crm_conn.commit()
        crm_conn.close()

        if added > 0:
            print(f"✅ naver-diagnosis에서 {added}개 업체 동기화 완료")
        return added

    except Exception as e:
        print(f"⚠️ 동기화 에러: {e}")
        return 0


# 앱 시작 시 동기화 — generate_first_message 정의 후 main에서 호출


# ──────────────────────────────────────────────────
# 파일명 정규화 함수
# ──────────────────────────────────────────────────

def extract_region_from_address(address: str) -> str:
    """주소에서 지역 추출 (예: "경기 양주시..." → "양주")"""
    if not address:
        return ""

    # "경기 양주시..." 형태에서 "양주" 추출
    parts = address.split()
    if len(parts) >= 2:
        region = parts[1]  # 두 번째 단어 (시/구 이름)
        # "양주시", "양주군" → "양주"로 정규화
        region = re.sub(r'(시|군|구)$', '', region)
        return region

    return ""

_PHONE_LOOKUP_CACHE = {}  # {업체명: (전화번호, 주소)}

def build_phone_lookup() -> dict:
    """
    소상공인 010수집 Excel 파일들을 스캔해서 업체명 → (전화번호, 주소) 캐시 구축
    """
    global _PHONE_LOOKUP_CACHE
    if _PHONE_LOOKUP_CACHE:
        return _PHONE_LOOKUP_CACHE

    def detect_phone_col(headers):
        """전화번호 컬럼 감지 — '010번호' 우선, 아니면 실제 값 확인"""
        # 1순위: '010번호' 정확 매칭
        for i, h in enumerate(headers):
            if h == '010번호' or h == '전화번호':
                return i
        # 2순위: '010' 포함하는 헤더
        for i, h in enumerate(headers):
            if '010' in h:
                return i
        # 3순위: '전화' 포함
        for i, h in enumerate(headers):
            if '전화' in h:
                return i
        return None

    def detect_name_col(headers):
        for i, h in enumerate(headers):
            if h in ('업체명', '상호명', '업체이름', '이름'):
                return i
        for i, h in enumerate(headers):
            if '업체' in h or '상호' in h:
                return i
        return None

    try:
        import openpyxl, glob as _glob
        # glob.escape로 [] 특수문자 이스케이프 처리
        escaped_base = _glob.escape(str(EXCEL_DB_FOLDER))
        xlsx_files = (
            list(_glob.glob(escaped_base + "/db/*.xlsx")) +
            list(_glob.glob(escaped_base + "/output/*.xlsx")) +
            list(_glob.glob(escaped_base + "/*.xlsx"))
        )

        lookup = {}
        for f in xlsx_files:
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                raw_headers = [c.value for c in next(ws.iter_rows())]
                headers = [str(h) if h else '' for h in raw_headers]

                name_col = detect_name_col(headers)
                phone_col = detect_phone_col(headers)
                addr_col = next((i for i, h in enumerate(headers) if '주소' in h), None)

                if name_col is None or phone_col is None:
                    wb.close()
                    continue

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) <= max(name_col, phone_col):
                        continue
                    name = str(row[name_col]).strip() if row[name_col] else ''
                    phone = str(row[phone_col]).strip() if row[phone_col] else ''
                    addr = str(row[addr_col]).strip() if addr_col is not None and row[addr_col] else ''

                    if name and phone and phone.startswith('010') and name not in lookup:
                        lookup[name] = (phone, addr)
                wb.close()
            except Exception:
                pass

        _PHONE_LOOKUP_CACHE = lookup
        print(f"📋 Excel 전화번호 캐시 구축: {len(lookup)}개 업체")
        return lookup
    except Exception as e:
        print(f"⚠️ Excel 전화번호 캐시 구축 실패: {e}")
        return {}


def get_phone_from_excel(business_name: str):
    """업체명으로 Excel에서 전화번호와 주소 조회. (phone, addr) 반환"""
    lookup = build_phone_lookup()
    if business_name in lookup:
        return lookup[business_name]
    # 유사 매칭 시도 (공백 차이 등)
    name_stripped = business_name.replace(' ', '')
    for key, val in lookup.items():
        if key.replace(' ', '') == name_stripped:
            return val
    return None, None


def update_crm_phones_from_excel():
    """CRM DB의 placeholder 전화번호를 Excel에서 찾아 업데이트"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        # 이미 실제 전화번호가 있는 것들 수집 (UNIQUE 충돌 방지)
        c.execute("SELECT phone FROM businesses WHERE phone NOT LIKE '진단%' AND phone != ''")
        used_phones = {r['phone'] for r in c.fetchall()}

        c.execute("SELECT id, name, phone, location FROM businesses WHERE phone LIKE '진단%' OR phone = '' OR phone IS NULL")
        rows = c.fetchall()

        updated = 0
        for row in rows:
            phone, addr = get_phone_from_excel(row['name'])
            if phone and phone not in used_phones:
                location = row['location'] or addr or ''
                c.execute("UPDATE businesses SET phone = ?, location = ? WHERE id = ?",
                          (phone, location, row['id']))
                used_phones.add(phone)
                updated += 1

        conn.commit()
        conn.close()
        if updated > 0:
            print(f"✅ CRM 전화번호 업데이트: {updated}개 업체")
        return updated
    except Exception as e:
        print(f"⚠️ CRM 전화번호 업데이트 실패: {e}")
        return 0


def normalize_filename(business_name: str, phone: str, region: str = "") -> str:
    """
    지역-업체명-전화번호 형태로 파일명 정규화
    전화번호는 하이픈 유지 (010-8870-1150)

    예: "하쿠나마타타헤어", "010-8870-1150", "양주" → "양주-하쿠나마타타헤어-010-8870-1150"
    """
    # 업체명에서 파일명에 쓸 수 없는 문자만 제거
    safe_name = re.sub(r'[\\/:*?"<>|]', '', business_name).strip()
    if not safe_name:
        safe_name = "업체"

    # 전화번호 정리 — "진단-xxx" placeholder는 무시, 하이픈 유지
    safe_phone = ''
    if phone and not phone.startswith('진단') and phone != 'unknown':
        safe_phone = phone  # 하이픈 유지

    # 지역이 없으면 환경변수에서 가져오기
    if not region:
        region = os.getenv("SALES_REGION", "")

    # 지역 정규화 (특수문자 제거)
    safe_region = re.sub(r'[\\/:*?"<>|]', '', region).strip() if region else ""

    # 파일명 조합 — 순서: 지역-업체명-전화번호
    parts = []
    if safe_region:
        parts.append(safe_region)
    parts.append(safe_name)
    if safe_phone:
        parts.append(safe_phone)

    return '-'.join(parts)

# ──────────────────────────────────────────────────
# PDF 감시 + 업체 추출
# ──────────────────────────────────────────────────

def extract_business_info(filename):
    """
    파일명에서 번호(010-xxxx)와 업체명 추출
    예: 010-1234_양주헤어림.pdf → ('010-1234', '양주헤어림')
    """
    name = Path(filename).stem  # 확장자 제거

    # 010으로 시작하는 번호 추출
    parts = name.split('_')
    if not parts:
        return None, None

    phone = parts[0]  # 010-xxxx
    business_name = '_'.join(parts[1:]) if len(parts) > 1 else '미등록'

    return phone, business_name

def monitor_pdfs():
    """백그라운드에서 PDF 폴더 감시"""
    monitored = set()

    while True:
        try:
            if PDF_FOLDER.exists():
                for pdf_file in PDF_FOLDER.glob("*.pdf"):
                    if pdf_file.name not in monitored:
                        phone, name = extract_business_info(pdf_file.name)

                        if phone:
                            add_business(phone, name, pdf_file.name)
                            monitored.add(pdf_file.name)
                            print(f"✅ 업체 추가: {phone} | {name}")
        except Exception as e:
            print(f"⚠️ PDF 감시 에러: {e}")

        # 30초마다 naver-diagnosis DB 동기화도 실행
        sync_from_naver_diagnosis()

        time.sleep(5)  # 5초마다 확인

def get_naver_diagnosis_by_name(business_name: str) -> dict:
    """
    naver-diagnosis DB에서 업체명으로 진단 데이터 조회
    place_url, ppt_filename, 점수 등을 반환

    매칭 전략:
    1. 정확 매칭: WHERE business_name = ? (완전한 데이터 우선 — competitor_avg_review > 0)
    2. 공백 제거 매칭: REPLACE(business_name, ' ', '') = REPLACE(?, ' ', '')
    3. 부분 매칭: WHERE business_name LIKE ?
    4. 지역명 제거 매칭: 업체명에서 지역명 제거 후 재매칭
    """
    try:
        if not NAVER_DIAGNOSIS_DB.exists():
            return None

        conn = sqlite3.connect(str(NAVER_DIAGNOSIS_DB))
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        row = None

        # 1단계: 정확 매칭 (완전한 데이터 우선 — 경쟁사 데이터 있는 것 먼저)
        c.execute("""
            SELECT id, place_url, ppt_filename, total_score, grade
            FROM diagnosis_history
            WHERE business_name = ? AND competitor_avg_review > 0
            ORDER BY created_at DESC
            LIMIT 1
        """, (business_name,))
        row = c.fetchone()

        # 1-1단계: 정확 매칭이지만 경쟁사 데이터 없으면 그냥 최신
        if not row:
            c.execute("""
                SELECT id, place_url, ppt_filename, total_score, grade
                FROM diagnosis_history
                WHERE business_name = ?
                ORDER BY created_at DESC
                LIMIT 1
            """, (business_name,))
            row = c.fetchone()

        # 2단계: 공백 제거 매칭 (1단계 실패 시) — 역시 완전한 데이터 우선
        if not row:
            c.execute("""
                SELECT id, place_url, ppt_filename, total_score, grade
                FROM diagnosis_history
                WHERE REPLACE(business_name, ' ', '') = REPLACE(?, ' ', '') AND competitor_avg_review > 0
                ORDER BY created_at DESC
                LIMIT 1
            """, (business_name,))
            row = c.fetchone()

        # 2-1단계: 공백 제거 매칭이지만 경쟁사 데이터 없으면 그냥 최신
        if not row:
            c.execute("""
                SELECT id, place_url, ppt_filename, total_score, grade
                FROM diagnosis_history
                WHERE REPLACE(business_name, ' ', '') = REPLACE(?, ' ', '')
                ORDER BY created_at DESC
                LIMIT 1
            """, (business_name,))
            row = c.fetchone()

        # 3단계: 부분 매칭 (1,2단계 실패 시) — 완전한 데이터 우선
        if not row:
            c.execute("""
                SELECT id, place_url, ppt_filename, total_score, grade
                FROM diagnosis_history
                WHERE business_name LIKE ? AND competitor_avg_review > 0
                ORDER BY created_at DESC
                LIMIT 1
            """, (f"%{business_name}%",))
            row = c.fetchone()

        # 3-1단계: 부분 매칭이지만 경쟁사 데이터 없으면 그냥 최신
        if not row:
            c.execute("""
                SELECT id, place_url, ppt_filename, total_score, grade
                FROM diagnosis_history
                WHERE business_name LIKE ?
                ORDER BY created_at DESC
                LIMIT 1
            """, (f"%{business_name}%",))
            row = c.fetchone()

        # 4단계: 지역명 제거 매칭 (1,2,3단계 실패 시) — 역시 완전한 데이터 우선
        if not row:
            regions = ['양주', '포천', '의정부', '동두천', '연천', '가평', '남양주']
            for region in regions:
                cleaned = business_name.replace(region, '').strip()
                if cleaned and cleaned != business_name:
                    c.execute("""
                        SELECT id, place_url, ppt_filename, total_score, grade
                        FROM diagnosis_history
                        WHERE business_name LIKE ? AND competitor_avg_review > 0
                        ORDER BY created_at DESC
                        LIMIT 1
                    """, (f'%{cleaned}%',))
                    row = c.fetchone()
                    if row:
                        break

        # 4-1단계: 지역명 제거 매칭이지만 경쟁사 데이터 없으면 그냥 최신
        if not row:
            regions = ['양주', '포천', '의정부', '동두천', '연천', '가평', '남양주']
            for region in regions:
                cleaned = business_name.replace(region, '').strip()
                if cleaned and cleaned != business_name:
                    c.execute("""
                        SELECT id, place_url, ppt_filename, total_score, grade
                        FROM diagnosis_history
                        WHERE business_name LIKE ?
                        ORDER BY created_at DESC
                        LIMIT 1
                    """, (f'%{cleaned}%',))
                    row = c.fetchone()
                    if row:
                        break

        conn.close()

        if not row:
            return None

        return dict(row)
    except Exception as e:
        print(f"⚠️ naver-diagnosis 조회 실패: {e}")
        return None


def add_business(phone, name, pdf_filename):
    """업체 DB에 추가"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        # 중복 체크
        c.execute("SELECT id FROM businesses WHERE phone = ?", (phone,))
        if c.fetchone():
            conn.close()
            return

        # naver-diagnosis에서 추가 정보 조회
        diagnosis_info = get_naver_diagnosis_by_name(name)
        place_url = diagnosis_info.get('place_url') if diagnosis_info else None
        naver_diagnosis_id = diagnosis_info.get('id') if diagnosis_info else None

        c.execute("""INSERT INTO businesses (phone, name, pdf_path, place_url, naver_diagnosis_id)
                     VALUES (?, ?, ?, ?, ?)""",
                 (phone, name, pdf_filename, place_url, naver_diagnosis_id))

        business_id = c.lastrowid

        # 1차 메시지 생성 후 저장
        first_message = generate_first_message(name, phone)

        # conversations 테이블에 저장 (호환성 유지)
        c.execute("""INSERT INTO conversations (business_id, speaker, message, message_order)
                     VALUES (?, ?, ?, ?)""", (business_id, 'AI', first_message, 1))

        # message_sequences 테이블에도 저장 (시퀀스 관리)
        c.execute("""INSERT INTO message_sequences (business_id, sequence_num, message)
                     VALUES (?, ?, ?)""", (business_id, 1, first_message))

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"⚠️ 업체 추가 에러: {e}")

# ──────────────────────────────────────────────────
# AI 메시지 생성
# ──────────────────────────────────────────────────

def generate_first_message(name, phone):
    """
    1차 메시지 생성
    naver-diagnosis DB에서 진단 데이터 조회 → message_generator 사용
    메시지 캐싱: ai_first_message 컬럼에서 기존 메시지 먼저 확인
    """
    try:
        # naver-diagnosis DB에서 진단 데이터 조회
        diagnosis_data = get_diagnosis_data(name)

        # DB에 없으면 기본 데이터 사용
        if not diagnosis_data:
            diagnosis_data = get_default_diagnosis_data(name, phone)

        # ── 캐싱 체크: ai_first_message 컬럼 확인 ────────────────
        if diagnosis_data and diagnosis_data.get('ai_first_message'):
            print(f"✅ 캐시된 메시지 사용: {name}")
            return diagnosis_data['ai_first_message']

        # diagnosis_data 품질 개선
        # category가 없거나 "기타"면 업체명에서 추론
        if not diagnosis_data.get('category') or diagnosis_data.get('category') == '기타':
            diagnosis_data['category'] = infer_category_from_name(name)

        # naver_place_rank가 0이거나 없으면 기본값 설정
        if not diagnosis_data.get('naver_place_rank') or diagnosis_data.get('naver_place_rank') <= 0:
            diagnosis_data['naver_place_rank'] = 15  # 기본 15위

        generated_message = None

        # message_generator가 available하면 사용
        if MESSAGE_GENERATOR_AVAILABLE:
            try:
                msg_result = msg_gen_first(diagnosis_data)
                # message_generator 반환 형식: {type, text, sms_text, label}
                generated_message = msg_result.get("text", "")
            except Exception as e:
                print(f"⚠️ message_generator 실행 실패: {e}. Gemini로 대체")

        # message_generator 없거나 실패 → Gemini 사용
        if not generated_message:
            rank = diagnosis_data.get('naver_place_rank', 15)
            my_review = diagnosis_data.get('review_count', 0)
            comp_review = diagnosis_data.get('competitor_avg_review', 0)
            category = diagnosis_data.get('category', '업종')
            region = os.environ.get('SALES_REGION', '')
            search_q = f"{region} {category}".strip() if region else category

            if comp_review > 0 and comp_review >= my_review * 2:
                prompt = f"""너는 네이버 플레이스 마케팅 영업사원이야. 카톡 첫 메시지를 써줘.

업체명: {name}
상황: 주변 상위 가게 리뷰 평균 {comp_review}개, {name}은 {my_review}개

규칙 (무조건):
- 마크다운 완전 금지. *, **, #, _ 절대 사용 금지.
- 줄바꿈만 사용.
- 5줄 이내.
- 자기소개 없음 ("저희는" 금지).
- "혹시 이 차이 눈에 띄신 적 있으세요?" 로 끝내기.
- 완성본만 출력 (설명 없이).
"""
            elif rank > 0:
                prompt = f"""너는 네이버 플레이스 마케팅 영업사원이야. 카톡 첫 메시지를 써줘.

업체명: {name}
상황: "{search_q}" 검색 시 {rank}위권, 최근 업데이트 부족

규칙 (무조건):
- 마크다운 완전 금지. *, **, #, _ 절대 사용 금지.
- 줄바꿈만 사용.
- 5줄 이내.
- 자기소개 없음.
- "혹시 이런 부분 신경 쓰고 계신 편인가요?" 로 끝내기.
- 완성본만 출력 (설명 없이).
"""
            else:
                prompt = f"""너는 네이버 플레이스 마케팅 영업사원이야. 카톡 첫 메시지를 써줘.

업체명: {name}
상황: 네이버 플레이스 관리 상태 미흡

규칙 (무조건):
- 마크다운 완전 금지. *, **, #, _ 절대 사용 금지.
- 줄바꿈만 사용.
- 5줄 이내.
- 자기소개 없음.
- 질문으로 끝내기.
- 완성본만 출력 (설명 없이).
"""
            model = genai.GenerativeModel('gemini-2.5-flash')
            response = model.generate_content(prompt)
            generated_message = response.text

        # ** 마크다운 제거 (어떤 경로로 생성됐든 무조건)
        if generated_message:
            import re as _re
            generated_message = _re.sub(r'\*\*(.+?)\*\*', r'\1', generated_message)
            generated_message = _re.sub(r'\*\*', '', generated_message)
            generated_message = _re.sub(r'\*(.+?)\*', r'\1', generated_message)
            generated_message = generated_message.strip()

        # ── 캐시 저장: naver-diagnosis DB ai_first_message 컬럼에 저장 ────
        if generated_message and NAVER_DIAGNOSIS_DB.exists():
            try:
                nc = sqlite3.connect(str(NAVER_DIAGNOSIS_DB))
                nc.execute(
                    "UPDATE diagnosis_history SET ai_first_message = ? WHERE business_name = ? AND ai_first_message IS NULL",
                    (generated_message, name)
                )
                nc.commit()
                nc.close()
                print(f"✅ 메시지 캐시 저장: {name}")
            except Exception as e:
                print(f"⚠️ 메시지 캐시 저장 실패: {e}")

        return generated_message if generated_message else f"{name} 사장님, 데이터 정리하다가 연락드렸습니다. 확인 가능하신가요?"

    except Exception as e:
        print(f"❌ 1차 메시지 생성 실패: {e}")
        # 최악의 경우 폴백 메시지
        return f"{name} 사장님, 데이터 정리하다가 연락드렸습니다. 확인 가능하신가요?"

def detect_rejection_type(customer_message: str) -> str:
    """
    고객 메시지에서 거절 유형 감지
    Returns: "비싸다" | "직접" | "보류" | "경험있음" | "기타"
    """
    msg_lower = customer_message.lower()

    # 비싸다 감지
    if any(word in msg_lower for word in ["비싸", "비용", "금액", "가격", "비쌈", "너무", "많아"]):
        return "비싸다"

    # 직접 감지
    if any(word in msg_lower for word in ["직접", "혼자", "제가 할", "우리가", "할 수", "스스로"]):
        return "직접"

    # 보류 감지
    if any(word in msg_lower for word in ["나중에", "생각해", "괜찮아", "됐어", "안 돼", "아직"]):
        return "보류"

    # 경험있음 감지
    if any(word in msg_lower for word in ["해봤", "효과", "없었", "전에", "이미", "안 됐"]):
        return "경험있음"

    # 무응답은 별도 처리 (여기서는 기타로)
    return "기타"


def generate_response_messages(name, phone, customer_message):
    """
    고객 응답에 대한 최강의 1가지 답변 생성
    1. 거절 유형 감지
    2. message_generator에서 해당 유형 메시지 가져오기 (있으면)
    3. 없으면 Gemini로 자유 대화 생성
    """
    try:
        # 거절 유형 감지
        rejection_type = detect_rejection_type(customer_message)
        print(f"📊 거절 유형 감지: {rejection_type}")

        # naver-diagnosis DB에서 진단 데이터 조회
        diagnosis_data = get_diagnosis_data(name)
        if not diagnosis_data:
            diagnosis_data = get_default_diagnosis_data(name, phone)

        # message_generator가 available하면 사용
        if MESSAGE_GENERATOR_AVAILABLE and rejection_type != "기타":
            try:
                fourth_messages = msg_gen_fourth(diagnosis_data)
                # fourth_messages는 dict: {보류, 무응답, 비싸다, 직접, 경험있음}
                if rejection_type in fourth_messages:
                    template_msg = fourth_messages[rejection_type]
                    if template_msg:
                        print(f"✅ message_generator {rejection_type} 메시지 사용")
                        return template_msg
            except Exception as e:
                print(f"⚠️ message_generator 4차 메시지 실행 실패: {e}")

        # message_generator 없거나 기타면 Gemini로 자유 대화
        print("📝 Gemini로 자유 대화 생성")
        prompt = f"""
너는 실전 영업의 대가야. 네이버 플레이스 마케팅 대행사의 베테랑 영업사원이고,
월 100명의 사장님을 설득해본 경험이 있어.

[업체 정보]
- 업체명: {name}
- 번호: {phone}
- 종합점수: {diagnosis_data.get('total_score', 0):.0f}점
- 등급: {diagnosis_data.get('grade', 'D')}

[고객이 보낸 메시지]
{customer_message}

## 미션
이 메시지에 대해서 사장님이 "그래, 한번 상담받아봐야겠다"고 생각하게 만드는
카톡 1개 메시지를 만들어줘.

## 절대 규칙
- 구체적 해결책 절대 금지 ("사진을 이렇게..." X)
- 월액만 언급. 총액이나 프리미엄 X
- 자기소개 금지 ("저희 회사는..." X)
- 기술 용어 금지 ("알고리즘, SEO..." X)

## 톤 전략
1. 논리: 수치/팩트로 문제의 심각성 강조
2. 공감: 사장님의 입장 이해하기
3. 긴급: "지금" 결정하면 되는 근거

## 출력 형식
메시지만 (설명/제목 없이). 카톡에 그대로 복붕 가능하도록 3~5줄.
"""
        model = genai.GenerativeModel('gemini-2.5-flash')
        response = model.generate_content(prompt)
        return response.text

    except Exception as e:
        print(f"❌ 응답 메시지 생성 실패: {e}")
        # 최악의 경우 폴백 메시지
        return f"그 말씀이 현실적인 얘기네요. 혹시 시간 10분만 내주셔서 계산 한번 확인해보실래요?"

# 앱 시작 시 전화번호 캐시 구축 + CRM DB 업데이트 (모든 함수 정의 후 실행)
try:
    build_phone_lookup()
    update_crm_phones_from_excel()
except Exception as _e:
    print(f"⚠️ 초기 전화번호 업데이트 스킵: {_e}")

# ──────────────────────────────────────────────────
# API 라우트
# ──────────────────────────────────────────────────

@app.route('/')
def dashboard():
    """대시보드 (업체 리스트)"""
    return render_template('dashboard.html')

@app.route('/api/businesses')
def get_businesses():
    """모든 업체 조회"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute("""SELECT id, phone, name, status, created_at, place_url
                 FROM businesses ORDER BY created_at DESC""")
    businesses = [dict(row) for row in c.fetchall()]

    conn.close()
    return jsonify(businesses)

@app.route('/business/<int:business_id>')
def chat_window(business_id):
    """개별 채팅창"""
    return render_template('chat.html', business_id=business_id)

@app.route('/api/business/<int:business_id>')
def get_business(business_id):
    """업체 상세 정보"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute("SELECT * FROM businesses WHERE id = ?", (business_id,))
    business = dict(c.fetchone() or {})

    conn.close()
    return jsonify(business)

@app.route('/api/business/<int:business_id>/chat')
def get_chat_history(business_id):
    """대화 기록 조회 — 비어있으면 1차 멘트 자동 생성"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute("""SELECT speaker, message, created_at
                 FROM conversations
                 WHERE business_id = ?
                 ORDER BY created_at ASC""", (business_id,))

    messages = [dict(row) for row in c.fetchall()]

    # 대화가 비어있으면 1차 멘트 자동 생성
    if not messages:
        c.execute("SELECT name, phone FROM businesses WHERE id = ?", (business_id,))
        biz = c.fetchone()
        if biz:
            first_msg = generate_first_message(biz['name'], biz['phone'] or '')
            if first_msg:
                c.execute("""INSERT INTO conversations (business_id, speaker, message)
                             VALUES (?, ?, ?)""", (business_id, 'AI', first_msg))
                conn.commit()
                messages = [{'speaker': 'AI', 'message': first_msg, 'created_at': datetime.now().isoformat()}]

    conn.close()
    return jsonify(messages)

@app.route('/api/business/<int:business_id>/respond', methods=['POST'])
def add_customer_response(business_id):
    """고객 응답 추가 + AI 메시지 생성"""
    data = request.json
    customer_message = data.get('message', '')

    if not customer_message:
        return jsonify({'error': '메시지가 비어있습니다'}), 400

    # 고객 응답 저장
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("SELECT name, phone, status FROM businesses WHERE id = ?", (business_id,))
    business = c.fetchone()
    if not business:
        return jsonify({'error': '업체를 찾을 수 없습니다'}), 404

    name, phone, current_status = business

    # 고객 메시지 저장
    c.execute("""INSERT INTO conversations (business_id, speaker, message)
                 VALUES (?, ?, ?)""", (business_id, '고객', customer_message))

    conn.commit()
    conn.close()

    # AI 응답 생성
    try:
        ai_response = generate_response_messages(name, phone, customer_message)

        # AI 응답 저장
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""INSERT INTO conversations (business_id, speaker, message)
                     VALUES (?, ?, ?)""", (business_id, 'AI', ai_response))

        # 상태 업데이트 로직
        status_map = {
            '1차_발송_대기': '1차_발송_완료',
            '1차_발송_완료': '2차_생성',
            '2차_생성': '2차_발송_대기',
            '2차_발송_대기': '2차_발송_완료',
            '2차_발송_완료': '3차_생성',
            '3차_생성': '3차_발송_대기',
            '3차_발송_대기': '3차_발송_완료',
            '3차_발송_완료': '4차_생성',
            '4차_생성': '4차_발송_대기',
        }

        next_status = status_map.get(current_status, current_status)
        c.execute("UPDATE businesses SET status = ? WHERE id = ?", (next_status, business_id))
        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'ai_response': ai_response,
            'next_status': next_status
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/business/<int:business_id>/report')
def view_html_report(business_id):
    """HTML 제안서를 브라우저에서 렌더링"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        c.execute("SELECT name, phone FROM businesses WHERE id = ?", (business_id,))
        business = c.fetchone()
        conn.close()

        if not business:
            return jsonify({'error': '업체를 찾을 수 없습니다'}), 404

        name = business['name']
        phone = business['phone']

        # naver-diagnosis DB에서 진단 데이터 조회
        diagnosis_data = get_diagnosis_data(name)
        if not diagnosis_data:
            diagnosis_data = get_default_diagnosis_data(name, phone)

        # HtmlPdfGenerator로 HTML 렌더링
        if not HTML_PDF_AVAILABLE:
            return jsonify({'error': 'HTML 생성 모듈을 찾을 수 없습니다'}), 500

        try:
            generator = HtmlPdfGenerator()
            html = generator.render_html(diagnosis_data)
            return html, 200, {'Content-Type': 'text/html; charset=utf-8'}
        except Exception as e:
            print(f"❌ HTML 렌더링 실패: {e}")
            return jsonify({'error': f'HTML 렌더링 실패: {str(e)}'}), 500

    except Exception as e:
        return jsonify({'error': str(e)}), 500


DESKTOP = os.path.expanduser("~/Desktop")
SALES_DIR = os.path.join(DESKTOP, "영업")


def _get_business_folder(name, phone, region):
    """바탕화면/영업/ 아래 지역-업체명-번호 폴더 생성 후 경로 반환"""
    folder_name = normalize_filename(name, phone, region)
    folder_path = os.path.join(SALES_DIR, folder_name)
    os.makedirs(folder_path, exist_ok=True)
    return folder_path, folder_name


def _get_business_data(business_id):
    """업체 정보 + 진단 데이터 + 지역 한번에 가져오기"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT name, phone FROM businesses WHERE id = ?", (business_id,))
    business = c.fetchone()
    if not business:
        conn.close()
        return None, None, None, None, None

    name = business['name']
    phone = business['phone'] or ''

    # phone이 placeholder면 같은 업체명의 다른 row에서 진짜 번호 찾기
    if not phone or phone.startswith('진단'):
        c.execute("SELECT phone FROM businesses WHERE name = ? AND phone NOT LIKE '진단%' AND phone != '' LIMIT 1", (name,))
        real = c.fetchone()
        if real and real['phone']:
            phone = real['phone']
        else:
            # Excel 파일에서 전화번호 조회
            excel_phone, excel_addr = get_phone_from_excel(name)
            if excel_phone:
                phone = excel_phone
                # CRM DB도 업데이트
                c.execute("UPDATE businesses SET phone = ?, location = ? WHERE id = ?",
                          (excel_phone, excel_addr or '', business_id))
                conn.commit()

    conn.close()

    diagnosis_data = get_diagnosis_data(name)
    if not diagnosis_data:
        diagnosis_data = get_default_diagnosis_data(name, phone)

    addr = diagnosis_data.get('address', '')
    region = extract_region_from_address(addr) if addr else os.environ.get('SALES_REGION', '')

    return name, phone, region, diagnosis_data, business


@app.route('/api/business/<int:business_id>/pdf')
def download_pdf(business_id):
    """PDF를 바탕화면 폴더에 저장 — Playwright로 HTML→PDF 변환"""
    try:
        name, phone, region, diagnosis_data, biz = _get_business_data(business_id)
        if not biz:
            return jsonify({'error': '업체를 찾을 수 없습니다'}), 404

        if not HTML_PDF_AVAILABLE:
            return jsonify({'error': 'HTML 생성 모듈 없음'}), 500

        # 1. HtmlPdfGenerator로 HTML 렌더링
        generator = HtmlPdfGenerator()
        html = generator.render_html(diagnosis_data)

        # 2. 바탕화면 폴더 생성
        folder_path, folder_name = _get_business_folder(name, phone, region)
        pdf_path = os.path.join(folder_path, f"{folder_name}.pdf")

        # 3. 임시 HTML 파일 저장
        tmp_html = os.path.join(folder_path, f"_tmp_{folder_name}.html")
        with open(tmp_html, 'w', encoding='utf-8') as f:
            f.write(html)

        # 4. Playwright로 PDF 변환 (naver-diagnosis venv 사용)
        NAVER_PYTHON = str(_SALES_TOOL_DIR / "naver-diagnosis" / "venv" / "Scripts" / "python.exe")

        # file:/// URI 만들기 (Windows 경로)
        file_uri = 'file:///' + tmp_html.replace('\\', '/')

        pdf_script = f'''
from playwright.sync_api import sync_playwright
with sync_playwright() as p:
    browser = p.chromium.launch()
    page = browser.new_page()
    page.goto("{file_uri}")
    page.wait_for_load_state("networkidle", timeout=15000)
    page.pdf(path=r"{pdf_path}", format="A4", print_background=True, margin={{"top": "0", "right": "0", "bottom": "0", "left": "0"}})
    browser.close()
'''

        import subprocess
        result = subprocess.run(
            [NAVER_PYTHON, '-c', pdf_script],
            capture_output=True, text=True, timeout=30
        )

        # 임시 HTML 삭제
        try:
            os.remove(tmp_html)
        except Exception:
            pass

        if result.returncode != 0:
            print(f"❌ Playwright PDF 실패: {result.stderr}")
            return jsonify({'error': f'PDF 생성 실패: {result.stderr[:200]}'}), 500

        # 성공
        return jsonify({
            'success': True,
            'message': '📋 PDF 저장 완료',
            'path': pdf_path,
            'folder': folder_path,
        })

    except Exception as e:
        print(f"❌ PDF 생성 실패: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/business/<int:business_id>/photo')
def get_summary_photo(business_id):
    """진단 요약 이미지를 바탕화면 폴더에 저장"""
    try:
        import datetime

        name, phone, region, diagnosis_data, biz = _get_business_data(business_id)
        if not biz:
            return jsonify({'error': '업체를 찾을 수 없습니다'}), 404

        # 발송 전 데이터 검증
        missing_fields = []
        if not diagnosis_data.get('naver_place_rank') or diagnosis_data.get('naver_place_rank') <= 0:
            missing_fields.append('순위')
        if not diagnosis_data.get('review_count') or diagnosis_data.get('review_count') <= 0:
            missing_fields.append('리뷰수')
        if missing_fields:
            print(f"⚠️ {name} — 일부 데이터 미수집: {missing_fields} (발송 가능하나 주의 필요)")

        # Pillow로 요약 이미지 생성
        from PIL import Image, ImageDraw, ImageFont
        from io import BytesIO

        grade = diagnosis_data.get('grade', 'D')
        total_score = diagnosis_data.get('total_score', 0)
        review_count = diagnosis_data.get('review_count', 0)
        photo_count = diagnosis_data.get('photo_count', 0)
        competitor_avg_review = diagnosis_data.get('competitor_avg_review', 0)
        competitor_avg_photo = diagnosis_data.get('competitor_avg_photo', 0)
        estimated_lost = diagnosis_data.get('estimated_lost_customers', 0)
        rank = diagnosis_data.get('naver_place_rank', 0)

        # 색상
        grade_colors = {'A': '#00c73c', 'B': '#4caf50', 'C': '#ff9800', 'D': '#f44336'}
        grade_color = grade_colors.get(grade, '#f44336')
        bg_color = '#1a1a2e'
        card_color = '#16213e'
        text_color = '#ffffff'
        sub_color = '#a0a0a0'
        accent = '#e94560'

        # 폰트 (시스템 폰트 사용)
        try:
            font_large = ImageFont.truetype("malgun.ttf", 36)
            font_medium = ImageFont.truetype("malgun.ttf", 24)
            font_small = ImageFont.truetype("malgun.ttf", 18)
            font_grade = ImageFont.truetype("malgunbd.ttf", 72)
        except Exception:
            font_large = ImageFont.load_default()
            font_medium = font_large
            font_small = font_large
            font_grade = font_large

        # 카드 구성 — 동적으로 필요한 카드만 포함
        cards = []

        # 현재 순위: 항상 표시 (0이면 "미확인")
        cards.append(("현재 순위", f"{rank}위" if rank > 0 else "미확인", "검색 결과 노출 위치"))

        # 종합 점수: 항상 표시
        cards.append(("종합 점수", f"{total_score:.0f}점", f"{grade}등급"))

        # 리뷰: review_count > 0일 때만
        if review_count > 0:
            desc = f"동네 평균 {competitor_avg_review}개" if competitor_avg_review > 0 else "리뷰 관리 필요"
            cards.append(("리뷰", f"{review_count}개", desc))

        # 사진: photo_count > 0일 때만, 0이면 "등록 필요"로 별도 표시
        if photo_count > 0:
            desc = f"동네 평균 {competitor_avg_photo}장" if competitor_avg_photo > 0 else "사진 관리 필요"
            cards.append(("사진", f"{photo_count}장", desc))
        else:
            cards.append(("사진", "등록 필요", "네이버 플레이스 사진 없음"))

        # 연간 기회손실: estimated_lost > 0일 때만
        if estimated_lost > 0:
            avg_price = 60000  # 업종별 기본 객단가
            annual_loss = estimated_lost * 12 * avg_price
            annual_loss_str = f"{annual_loss // 10000}만원" if annual_loss >= 10000 else f"{annual_loss:,}원"
            cards.append(("연간 기회손실", annual_loss_str, f"월 {estimated_lost}명 이탈 기준"))

        # 이미지 높이 동적 계산: 헤더(160) + 각 카드(140씩) + 하단 여백(100)
        card_height = len(cards) * 140
        H = 160 + card_height + 100
        W = 800

        img = Image.new('RGB', (W, H), bg_color)
        draw = ImageDraw.Draw(img)

        # 헤더
        draw.text((40, 30), "네이버 플레이스 진단", fill=sub_color, font=font_small)
        draw.text((40, 60), name, fill=text_color, font=font_large)

        # 등급 원 (오른쪽)
        cx, cy, r = 680, 80, 50
        draw.ellipse([cx-r, cy-r, cx+r, cy+r], fill=grade_color)
        bbox = draw.textbbox((0, 0), grade, font=font_grade)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        draw.text((cx - tw//2, cy - th//2 - 5), grade, fill='#fff', font=font_grade)

        # 구분선
        draw.line([(40, 140), (W-40, 140)], fill='#333', width=2)

        # 카드들 렌더링
        y = 170
        for label, value, desc in cards:
            # 카드 배경
            draw.rounded_rectangle([40, y, W-40, y+120], radius=12, fill=card_color)
            draw.text((70, y+15), label, fill=sub_color, font=font_small)
            draw.text((70, y+45), value, fill=accent, font=font_large)
            draw.text((70, y+88), desc, fill=sub_color, font=font_small)
            y += 140

        # 하단 워터마크 + 진단 정보
        diag_date = diagnosis_data.get('created_at', '')[:10] if diagnosis_data.get('created_at') else datetime.date.today().isoformat()
        diag_id = f"DIAG-{diag_date.replace('-','')}-{business_id:04d}"

        draw.text((40, H-65), f"진단번호: {diag_id}", fill='#555', font=font_small)
        draw.text((40, H-40), f"수집일: {diag_date}  |  리안 컴퍼니", fill='#444', font=font_small)

        # 바탕화면 폴더에 저장
        folder_path, folder_name = _get_business_folder(name, phone, region)
        filename = f"{folder_name}_진단요약.png"
        filepath = os.path.join(folder_path, filename)
        img.save(filepath, format='PNG', quality=95)

        return jsonify({
            'success': True,
            'message': f'📸 진단 사진 저장 완료',
            'path': filepath,
            'folder': folder_path,
        })

    except Exception as e:
        print(f"❌ 이미지 생성 실패: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/business/<int:business_id>/place')
def get_place_url(business_id):
    """네이버 플레이스 URL 조회"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute("SELECT name FROM businesses WHERE id = ?", (business_id,))
        business = c.fetchone()
        conn.close()

        if not business:
            return jsonify({'error': '업체를 찾을 수 없습니다'}), 404

        name = business[0]

        # naver-diagnosis DB에서 place_url 조회 시도
        place_url = None
        try:
            if NAVER_DIAGNOSIS_DB.exists():
                nconn = sqlite3.connect(str(NAVER_DIAGNOSIS_DB))
                nc = nconn.cursor()
                nc.execute("SELECT place_url FROM diagnosis_history WHERE business_name = ? ORDER BY created_at DESC LIMIT 1", (name,))
                row = nc.fetchone()
                if row and row[0]:
                    place_url = row[0]
                nconn.close()
        except Exception:
            pass

        # 없으면 네이버 검색 URL
        if not place_url:
            place_url = f"https://map.naver.com/p/search/{name}"

        return jsonify({'place_url': place_url})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/business/<int:business_id>/mark-sent', methods=['POST'])
def mark_sent(business_id):
    """멘트 발송 완료 표시"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute("SELECT status FROM businesses WHERE id = ?", (business_id,))
        business = c.fetchone()
        if not business:
            return jsonify({'error': '업체를 찾을 수 없습니다'}), 404

        current_status = business[0]

        # 상태 매핑: 대기 상태 → 완료 상태
        sent_status_map = {
            '1차_발송_대기': '1차_발송_완료',
            '2차_발송_대기': '2차_발송_완료',
            '3차_발송_대기': '3차_발송_완료',
            '4차_발송_대기': '계약완료',
        }

        next_status = sent_status_map.get(current_status)
        if not next_status:
            return jsonify({'error': f'이 상태에서는 발송 완료 처리할 수 없습니다: {current_status}'}), 400

        c.execute("UPDATE businesses SET status = ? WHERE id = ?", (next_status, business_id))
        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'next_status': next_status
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/business/<int:business_id>/next-message')
def get_next_message(business_id):
    """현재 상태에 맞는 다음 메시지 반환"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute("SELECT name, phone, status FROM businesses WHERE id = ?", (business_id,))
        business = c.fetchone()
        conn.close()

        if not business:
            return jsonify({'error': '업체를 찾을 수 없습니다'}), 404

        name, phone, current_status = business

        # 상태별로 생성할 메시지 번호 결정
        status_to_message_map = {
            '2차_생성': 2,
            '3차_생성': 3,
            '4차_생성': 4,
        }

        message_num = status_to_message_map.get(current_status)
        if not message_num:
            return jsonify({
                'success': True,
                'message': None,
                'info': '생성할 메시지가 없습니다. 고객 응답을 기다려주세요.'
            })

        # 이미 생성된 메시지가 있는지 확인
        c = conn.cursor()
        c.execute("""SELECT message FROM message_sequences
                     WHERE business_id = ? AND sequence_num = ?""",
                 (business_id, message_num))
        existing = c.fetchone()

        if existing:
            return jsonify({
                'success': True,
                'message': existing[0],
                'message_num': message_num
            })

        # 새 메시지 생성
        ai_response = generate_response_messages(name, phone, "")

        # DB에 저장
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""INSERT INTO message_sequences (business_id, sequence_num, message)
                     VALUES (?, ?, ?)""", (business_id, message_num, ai_response))
        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': ai_response,
            'message_num': message_num
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ──────────────────────────────────────────────────
# 백그라운드 스레드 시작
# ──────────────────────────────────────────────────

def start_pdf_monitor():
    """PDF 감시 스레드 시작"""
    thread = threading.Thread(target=monitor_pdfs, daemon=True)
    thread.start()

if __name__ == '__main__':
    print("\n" + "="*60)
    print("🚀 영업 CRM 시스템 시작")
    print("="*60)
    print("\n📍 접속: http://localhost:5000")
    print("\n📂 PDF 감시 폴더:")
    print(f"   {PDF_FOLDER}")
    print("\n💾 DB 위치:")
    print(f"   {DB_PATH}\n")

    # naver-diagnosis DB에서 업체 동기화
    sync_from_naver_diagnosis()

    start_pdf_monitor()

    # 브라우저 자동 열기 (1초 후)
    def open_browser():
        time.sleep(1)
        try:
            webbrowser.open('http://localhost:5000')
        except:
            pass

    threading.Thread(target=open_browser, daemon=True).start()

    app.run(debug=False, port=5000, use_reloader=False)
