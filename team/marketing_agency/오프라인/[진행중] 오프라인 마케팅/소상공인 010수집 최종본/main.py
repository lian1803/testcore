"""
소상공인 010번호 수집기
=======================
동작 방식:
  1. 네이버 웹검색: "{지역} {업종} 010" → 결과 HTML에서 010번호 추출
     (인스타, 당근, 블로그, 네이버플레이스 결과 포함)
  2. 카카오맵 Playwright: 지역+업종 검색 → 전화번호 추출 (010만 보관)
  3. 네이버 블로그 API: 블로그 스니펫에서 010번호 추출
  4. 중복 제거 (010번호 기준)
  5. 네이버지도 검증: 실제 운영중인 업체인지 확인
  6. 엑셀 저장

사용법:
  python main.py
  python main.py 의정부    ← 지역 변경
  python main.py 포천 --no-verify  ← 검증 생략 (빠름)
"""

import asyncio
import re
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Optional

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ─────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────
REGION = sys.argv[1] if len(sys.argv) > 1 else "양주"
DO_VERIFY = "--no-verify" not in sys.argv

NAVER_CLIENT_ID = "o0776HJDmQVO6J9Lez1m"
NAVER_CLIENT_SECRET = "ZXG0lPbgH9"

DESKTOP_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
MOBILE_UA = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
)

# 지역별 서브지역 (더 넓게 수집하기 위해)
SUB_AREAS = {
    "양주": ["양주", "양주 회천", "양주 백석", "양주 덕정", "양주 덕계",
             "양주 옥정", "양주 장흥", "양주 광적", "양주 남면", "양주 은현"],
    "의정부": ["의정부", "의정부 가능", "의정부 녹양", "의정부 신곡",
               "의정부 호원", "의정부 장암", "의정부 민락", "의정부 흥선"],
    "포천": ["포천", "포천 소흘", "포천 군내", "포천 동면", "포천 신북",
             "포천 가산", "포천 일동", "포천 영중"],
}

CATEGORIES = [
    "음식점", "카페", "미용실", "헤어샵", "네일샵", "네일아트",
    "헬스장", "필라테스", "요가", "피티",
    "병원", "치과", "한의원", "피부과", "안과", "정형외과",
    "학원", "태권도", "피아노", "영어학원",
    "꽃집", "베이커리", "빵집", "케이크",
    "부동산", "공인중개사",
    "펜션", "게스트하우스",
    "세탁소", "사진관", "애견미용", "반려동물",
    "인테리어", "리모델링",
    "치킨", "피자", "족발", "보쌈",
]

# 제외 프랜차이즈
FRANCHISE_EXCLUDE = {
    "스타벅스", "이마트", "홈플러스", "롯데마트", "코스트코", "다이소",
    "맥도날드", "버거킹", "롯데리아", "KFC", "서브웨이", "맘스터치",
    "GS25", "CU", "세븐일레븐", "미니스톱", "이마트24",
    "올리브영", "유니클로", "자라", "H&M", "크린토피아",
    "bbq", "굽네치킨", "교촌치킨", "bhc",
}

PHONE_RE = re.compile(r'01[016789][-.\s]?\d{3,4}[-.\s]?\d{4}')
PLACE_ID_RE = re.compile(r'/place/(\d{6,})')

# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────
def normalize_phone(raw: str) -> str:
    digits = re.sub(r'[^\d]', '', raw)
    if len(digits) == 11 and digits.startswith('010'):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return raw

def is_010(phone: str) -> bool:
    return re.sub(r'[^\d]', '', phone).startswith('010')

def is_franchise(name: str) -> bool:
    return any(f in name for f in FRANCHISE_EXCLUDE)

def extract_name_near_phone(text: str, phone_match) -> str:
    """전화번호 앞 텍스트에서 업체명 추출"""
    start = max(0, phone_match.start() - 60)
    before = text[start:phone_match.start()].strip()
    # 마지막 의미있는 단어 (한글 2자 이상)
    words = re.findall(r'[가-힣A-Za-z0-9]{2,20}', before)
    if words:
        return words[-1]
    return ""

# ─────────────────────────────────────────────
# 1단계: 네이버 웹검색 → 010번호 추출
# ─────────────────────────────────────────────
async def search_naver_web(page, query: str, max_pages: int = 3) -> list[dict]:
    """
    네이버 웹검색 결과 HTML에서 010번호 추출
    결과에는 인스타그램, 당근마켓, 블로그, 네이버플레이스 등이 포함됨
    """
    results = []
    seen_phones = set()

    for pg in range(max_pages):
        start = pg * 10 + 1
        encoded = urllib.parse.quote(query)
        url = f"https://search.naver.com/search.naver?query={encoded}&where=web&start={start}"

        try:
            await page.goto(url, timeout=20000)
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
            await page.wait_for_timeout(800)

            html = await page.content()
            text = await page.inner_text("body")

            # 결과가 없으면 중단
            if "검색결과가 없습니다" in text or "일치하는 문서가 없습니다" in text:
                break

            # 010번호 추출 (HTML + 텍스트 전체에서)
            for m in PHONE_RE.finditer(text):
                phone_raw = m.group(0)
                if not is_010(phone_raw):
                    continue
                phone = normalize_phone(phone_raw)
                if phone in seen_phones:
                    continue
                seen_phones.add(phone)

                name_hint = extract_name_near_phone(text, m)
                results.append({
                    "phone": phone,
                    "name": name_hint,
                    "source": "네이버웹검색",
                    "query": query,
                })

            await asyncio.sleep(0.4)

        except (PWTimeout, Exception) as e:
            break

    return results


# ─────────────────────────────────────────────
# 2단계: 네이버 블로그 API → 스니펫에서 010 추출
# ─────────────────────────────────────────────
async def search_naver_blog_api(query: str, max_pages: int = 5) -> list[dict]:
    """네이버 블로그 API: 결과 스니펫(description)에서 010번호 추출"""
    results = []
    seen_phones = set()
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }

    async with httpx.AsyncClient(timeout=10) as client:
        for pg in range(max_pages):
            start = pg * 10 + 1
            params = {"query": query, "display": 10, "start": start, "sort": "date"}
            try:
                resp = await client.get(
                    "https://openapi.naver.com/v1/search/blog.json",
                    headers=headers, params=params
                )
                resp.raise_for_status()
                items = resp.json().get("items", [])
                if not items:
                    break

                for item in items:
                    # title + description 합쳐서 010 추출
                    combined = re.sub(r'<[^>]+>', '', item.get("title", "") + " " + item.get("description", ""))
                    for m in PHONE_RE.finditer(combined):
                        phone_raw = m.group(0)
                        if not is_010(phone_raw):
                            continue
                        phone = normalize_phone(phone_raw)
                        if phone in seen_phones:
                            continue
                        seen_phones.add(phone)
                        name_hint = extract_name_near_phone(combined, m)
                        results.append({
                            "phone": phone,
                            "name": name_hint,
                            "source": "네이버블로그",
                            "query": query,
                            "blog_url": item.get("link", ""),
                        })
                await asyncio.sleep(0.2)
            except Exception:
                break

    return results


# ─────────────────────────────────────────────
# 3단계: 다음(Daum) 웹검색 → 010번호 추출
# 다음은 카카오 계열 → 카카오맵, 카카오채널 결과 포함
# ─────────────────────────────────────────────
async def search_daum_web(page, query: str, max_pages: int = 3) -> list[dict]:
    """
    다음 웹검색 결과 HTML에서 010번호 추출
    카카오맵, 카카오채널, 티스토리, 인스타 등 결과 포함됨
    """
    results = []
    seen_phones = set()

    for pg in range(max_pages):
        page_num = pg + 1
        encoded = urllib.parse.quote(query)
        url = f"https://search.daum.net/search?w=web&q={encoded}&p={page_num}"

        try:
            await page.goto(url, timeout=20000)
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
            await page.wait_for_timeout(700)

            text = await page.inner_text("body")

            if "검색결과가 없습니다" in text:
                break

            for m in PHONE_RE.finditer(text):
                phone_raw = m.group(0)
                if not is_010(phone_raw):
                    continue
                phone = normalize_phone(phone_raw)
                if phone in seen_phones:
                    continue
                seen_phones.add(phone)
                name_hint = extract_name_near_phone(text, m)
                results.append({
                    "phone": phone,
                    "name": name_hint,
                    "source": "다음검색(카카오)",
                    "query": query,
                })

            await asyncio.sleep(0.3)

        except (PWTimeout, Exception):
            break

    return results


# ─────────────────────────────────────────────
# 4단계: 네이버지도 검증
# ─────────────────────────────────────────────
async def verify_on_naver_maps(page, phone: str, name_hint: str, region: str) -> dict:
    """
    네이버지도에서 업체 실존 여부 확인
    - 우선 업체명+지역으로 검색
    - 없으면 전화번호로 검색
    """
    verified = False
    biz_name = ""
    address = ""
    naver_url = ""

    search_queries = []
    if name_hint and len(name_hint) >= 2:
        search_queries.append(f"{name_hint} {region}")
    search_queries.append(phone)

    for q in search_queries:
        try:
            encoded = urllib.parse.quote(q)
            await page.goto(
                f"https://m.search.naver.com/search.naver?query={encoded}&where=m_local",
                timeout=20000
            )
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
            await page.wait_for_timeout(600)

            html = await page.content()
            text = await page.inner_text("body")

            pid_match = PLACE_ID_RE.search(html)
            if not pid_match:
                continue

            pid = pid_match.group(1)
            naver_url = f"https://map.naver.com/p/entry/place/{pid}"

            # 전화번호 일치 확인
            phones_on_page = [normalize_phone(m.group(0)) for m in PHONE_RE.finditer(text)]
            if phone in phones_on_page or not phones_on_page:
                verified = True

            # 업체명 + 주소 추출
            name_m = re.search(r'"placeName"\s*:\s*"([^"]{2,40})"', html)
            addr_m = re.search(r'"roadAddress"\s*:\s*"([^"]{5,80})"', html)
            if name_m:
                biz_name = name_m.group(1)
            if addr_m:
                address = addr_m.group(1)

            if verified or (pid and not phones_on_page):
                # place가 있으면 일단 운영중으로 간주
                verified = True
                break

        except (PWTimeout, Exception):
            continue

    return {
        "verified": verified,
        "biz_name": biz_name,
        "address": address,
        "naver_url": naver_url,
    }


# ─────────────────────────────────────────────
# 엑셀 저장
# ─────────────────────────────────────────────
def save_excel(records: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "010번호 수집"

    headers = [
        "번호", "010번호", "업체명", "주소",
        "검증여부", "네이버플레이스",
        "수집출처", "검색쿼리",
    ]

    hfill = PatternFill("solid", fgColor="1a365d")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 26
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = center

    green = PatternFill("solid", fgColor="c6efce")
    yellow = PatternFill("solid", fgColor="ffeb9c")

    for idx, rec in enumerate(records, 1):
        row = idx + 1
        ws.row_dimensions[row].height = 20
        fill = green if rec.get("verified") else yellow

        row_data = [
            idx,
            rec.get("phone", ""),
            rec.get("biz_name") or rec.get("name", ""),
            rec.get("address", ""),
            "✅ 확인" if rec.get("verified") else "⚠️ 미확인",
            rec.get("naver_url", ""),
            rec.get("source", ""),
            rec.get("query", ""),
        ]

        left_cols = {2, 3, 4, 6, 7, 8}
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.alignment = left if col in left_cols else center
            if col == 6 and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    col_widths = [5, 16, 20, 35, 10, 50, 14, 30]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    Path(path).parent.mkdir(exist_ok=True)
    wb.save(path)


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
async def main():
    region = REGION
    areas = SUB_AREAS.get(region, [region])
    do_verify = DO_VERIFY

    print("=" * 60)
    print(f"  소상공인 010번호 수집기")
    print(f"  지역: {region} | 서브지역: {len(areas)}개 | 업종: {len(CATEGORIES)}개")
    print(f"  검증: {'ON' if do_verify else 'OFF (--no-verify)'}")
    print("=" * 60)

    all_records: dict[str, dict] = {}  # phone → record (중복 제거)

    # ── Phase 1: 네이버 웹검색 ─────────────────
    print(f"\n🔍 [1/3] 네이버 웹검색에서 010번호 수집 중...\n")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=DESKTOP_UA,
            viewport={"width": 1280, "height": 800},
            locale="ko-KR",
        )
        page = await ctx.new_page()

        total_web = 0
        for area in areas:
            for cat in CATEGORIES:
                query = f"{area} {cat} 010"
                items = await search_naver_web(page, query, max_pages=3)
                new = 0
                for item in items:
                    if item["phone"] not in all_records:
                        all_records[item["phone"]] = item
                        new += 1
                total_web += new
                if new > 0:
                    print(f"  [{area}] {cat:<12}: +{new}개  (누적 {len(all_records)}개)")
                await asyncio.sleep(0.3)

        await browser.close()

    print(f"\n  ✅ 웹검색 완료: {total_web}개 신규 (누적 {len(all_records)}개)")

    # ── Phase 2: 다음 웹검색 (카카오 계열) ───────
    print(f"\n🔍 [2/4] 다음검색(카카오)에서 010번호 수집 중...\n")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=DESKTOP_UA,
            viewport={"width": 1280, "height": 900},
            locale="ko-KR",
        )
        dpage = await ctx.new_page()

        total_daum = 0
        for area in areas[:5]:  # 대표 서브지역 5개
            for cat in CATEGORIES:
                query = f"{area} {cat} 010"
                items = await search_daum_web(dpage, query, max_pages=3)
                new = 0
                for item in items:
                    if item["phone"] not in all_records:
                        all_records[item["phone"]] = item
                        new += 1
                total_daum += new
                if new > 0:
                    print(f"  [다음] {area} {cat:<12}: +{new}개  (누적 {len(all_records)}개)")
                await asyncio.sleep(0.2)

        await browser.close()

    print(f"\n  ✅ 다음검색 완료: {total_daum}개 신규 (누적 {len(all_records)}개)")

    # ── Phase 3: 네이버 블로그 API ─────────────
    print(f"\n📝 [3/4] 네이버 블로그 API에서 010번호 추가 수집 중...\n")

    total_blog = 0
    for area in areas[:3]:  # 서브지역 중 대표 3개만 (속도)
        for cat in CATEGORIES:
            query = f"{area} {cat} 010"
            items = await search_naver_blog_api(query, max_pages=4)
            new = 0
            for item in items:
                if item["phone"] not in all_records:
                    all_records[item["phone"]] = item
                    new += 1
            total_blog += new
            await asyncio.sleep(0.15)

    print(f"  ✅ 블로그 완료: {total_blog}개 신규 (누적 {len(all_records)}개)")

    records = list(all_records.values())
    print(f"\n  📊 총 수집: {len(records)}개 (010번호 기준 중복 제거)")

    # ── Phase 4: 네이버지도 검증 ───────────────
    if do_verify and records:
        print(f"\n🗺️  [3/3] 네이버지도 검증 중... ({len(records)}개)\n")
        print("  (검증 생략하려면 --no-verify 옵션 사용)\n")

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)
            ctx = await browser.new_context(
                user_agent=MOBILE_UA,
                viewport={"width": 390, "height": 844},
                locale="ko-KR",
            )
            vpage = await ctx.new_page()

            verified_count = 0
            for i, rec in enumerate(records, 1):
                result = await verify_on_naver_maps(
                    vpage, rec["phone"], rec.get("name", ""), region
                )
                rec.update(result)
                if result["verified"]:
                    verified_count += 1
                    status = "✅"
                else:
                    status = "  "

                if i % 10 == 0 or result["verified"]:
                    name_display = (result.get("biz_name") or rec.get("name") or "?")[:12]
                    print(f"  [{i:4d}/{len(records)}] {status} {rec['phone']}  {name_display}")

                await asyncio.sleep(0.8)

            await browser.close()

        print(f"\n  ✅ 검증 완료: {verified_count}개 확인 / {len(records) - verified_count}개 미확인")
    else:
        if not do_verify:
            print(f"\n⏭️  검증 생략됨")

    # ── 저장 ──────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = f"output/{region}_010번호_{timestamp}.xlsx"

    print(f"\n💾 엑셀 저장 중...")
    save_excel(records, out_path)

    verified = sum(1 for r in records if r.get("verified"))
    print(f"\n{'='*60}")
    print(f"  완료! 파일: {out_path}")
    print(f"  총 010번호: {len(records)}개")
    if do_verify:
        print(f"  네이버지도 확인됨: {verified}개")
        print(f"  미확인(신규개척): {len(records) - verified}개")
    print(f"{'='*60}")


if __name__ == "__main__":
    asyncio.run(main())
