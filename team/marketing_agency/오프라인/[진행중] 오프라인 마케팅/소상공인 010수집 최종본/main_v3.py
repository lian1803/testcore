"""
소상공인 010번호 수집기 v3
============================
원본 TS 코드 기반 재구현 (3/23 삭제된 버전 복원)

소스:
  1. 카카오맵 API  — search.map.kakao.com/mapsearch/map.daum (JSONP, HTTP 직접, Playwright 불필요)
  2. 당근마켓      — /local-profile/ 페이지 방문
  3. DuckDuckGo   — html.duckduckgo.com/html/ (HTML버전, 가벼움)
  4. 네이버 웹검색 — search.naver.com 병렬 Playwright

사용법:
  python main_v3.py            ← 양주
  python main_v3.py 의정부
  python main_v3.py 포천 --verify
"""

import asyncio
import re
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ─────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────
REGION    = sys.argv[1] if len(sys.argv) > 1 else "양주"
DO_VERIFY = "--verify" in sys.argv

DESKTOP_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)
MOBILE_UA = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
)

# 카카오맵 카테고리 (원본 TS와 동일)
KAKAO_CATS = [
    # 음식점
    "치킨", "피자", "족발", "삼겹살", "분식", "중국집", "일식", "돈까스",
    "국밥", "냉면", "빵집", "도시락", "맛집", "식당", "떡볶이", "순대국",
    # 뷰티/헬스
    "미용실", "네일", "헬스장", "필라테스", "피부관리", "왁싱", "속눈썹",
    # 카페
    "카페", "커피숍", "버블티",
    # 교육
    "학원", "공부방", "과외", "레슨",
    # 생활
    "세탁소", "청소업체", "인테리어", "수리", "이사", "꽃집",
    # 반려동물
    "반려동물", "동물병원", "애견미용",
]

# 당근/DDG/네이버 카테고리
WEB_CATS = [
    "미용실", "카페", "맛집", "학원", "네일", "헬스", "필라테스",
    "인테리어", "세탁소", "꽃집", "청소", "과외", "레슨",
    "수리", "반려동물", "이사", "용달", "치킨", "피자", "족발",
]

FRANCHISE_EXCLUDE = {
    "스타벅스", "이마트", "홈플러스", "롯데마트", "코스트코", "다이소",
    "맥도날드", "버거킹", "롯데리아", "KFC", "서브웨이", "맘스터치",
    "GS25", "CU", "세븐일레븐", "미니스톱", "이마트24",
    "올리브영", "유니클로", "자라", "H&M", "크린토피아",
    "bbq", "굽네치킨", "교촌치킨", "bhc",
}

PHONE_RE    = re.compile(r'01[016789][-.\s]?\d{3,4}[-.\s]?\d{4}')
PLACE_ID_RE = re.compile(r'/place/(\d{6,})')

_lock   = asyncio.Lock()
_records: dict[str, dict] = {}

# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────
def normalize_phone(raw: str) -> str:
    d = re.sub(r'[^\d]', '', raw)
    if len(d) == 11: return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if len(d) == 10: return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    return raw

def is_010(raw: str) -> bool:
    d = re.sub(r'[^\d]', '', raw)
    return d.startswith('010') and len(d) == 11

def is_franchise(name: str) -> bool:
    return any(f in name for f in FRANCHISE_EXCLUDE)

async def add_record(phone_raw: str, name: str, source: str, category: str = ""):
    if not is_010(phone_raw):
        return
    phone = normalize_phone(phone_raw)
    if is_franchise(name):
        return
    async with _lock:
        if phone not in _records:
            _records[phone] = {
                "phone": phone, "name": name, "category": category,
                "source": source, "verified": False,
                "biz_name": "", "address": "", "naver_url": "",
            }


# ─────────────────────────────────────────────
# Phase 1: 카카오맵 API (핵심 소스)
# ─────────────────────────────────────────────
async def kakao_search_page(client: httpx.AsyncClient, query: str, page: int) -> list:
    """JSONP 엔드포인트 직접 호출"""
    url = (
        "https://search.map.kakao.com/mapsearch/map.daum"
        f"?callback=cb&q={urllib.parse.quote(query)}&page={page}&size=15&sort=0"
    )
    headers = {
        "User-Agent": DESKTOP_UA,
        "Referer": "https://map.kakao.com/",
        "Accept-Language": "ko-KR,ko;q=0.9",
    }
    try:
        resp = await client.get(url, headers=headers, timeout=10)
        text = resp.text
        # JSONP 래퍼 제거: /**/cb({...});
        import json
        json_str = re.sub(r'^/\*\*/cb\(', '', text.strip())
        json_str = re.sub(r'\);\s*$', '', json_str)
        parsed = json.loads(json_str)
        return parsed.get("place", [])
    except Exception:
        return []


async def phase_kakao(region: str) -> int:
    short = region.replace("경기도", "").replace("시", "").replace("군", "").strip()
    new_count = 0

    async with httpx.AsyncClient() as client:
        for cat in KAKAO_CATS:
            query = f"{short} {cat}"
            for pg in range(1, 6):  # 최대 5페이지 × 15개 = 75개/카테고리
                places = await kakao_search_page(client, query, pg)
                if not places:
                    break
                for place in places:
                    tel = place.get("tel", "")
                    name = place.get("name", "")
                    if is_010(tel):
                        before = len(_records)
                        await add_record(tel, name, "카카오맵", cat)
                        if len(_records) > before:
                            new_count += 1
                await asyncio.sleep(0.3)
            await asyncio.sleep(0.5)

    return new_count


# ─────────────────────────────────────────────
# Phase 2: 당근마켓 (local-profile 페이지)
# ─────────────────────────────────────────────
async def daangn_worker(page, categories: list, short_region: str) -> int:
    new_count = 0
    for cat in categories:
        q = f"{short_region} {cat}"
        search_url = f"https://www.daangn.com/kr/local-profile/s/?search={urllib.parse.quote(q)}"
        try:
            await page.goto(search_url, timeout=25000)
            await page.wait_for_load_state("networkidle", timeout=20000)
            await asyncio.sleep(1.5)

            # 업체 프로필 링크 수집
            links = await page.evaluate("""() => {
                return Array.from(document.querySelectorAll('a[href*="/local-profile/"]'))
                    .map(a => a.href)
                    .filter(h => !/\\/local-profile\\/s\\//.test(h))
                    .filter((v,i,arr) => arr.indexOf(v) === i)
                    .slice(0, 10)
            }""")

            for link in links:
                try:
                    await page.goto(link, timeout=18000)
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                    await asyncio.sleep(1.0)

                    data = await page.evaluate("""() => {
                        const h = document.querySelector('h1, h2')?.textContent?.trim() || ''
                        const body = document.body.innerText
                        const phoneMatches = body.match(/010[-\\s]?\\d{3,4}[-\\s]?\\d{4}/g) || []
                        const telLinks = Array.from(document.querySelectorAll('a[href^="tel:"]'))
                            .map(a => a.href.replace('tel:', '').trim())
                        return { bizName: h, phones: phoneMatches, tel: telLinks }
                    }""")

                    all_phones = list(set(
                        data.get("phones", []) +
                        [t for t in data.get("tel", []) if is_010(t)]
                    ))
                    for phone in all_phones:
                        before = len(_records)
                        await add_record(phone, data.get("bizName", cat), "당근마켓", cat)
                        if len(_records) > before:
                            new_count += 1
                except Exception:
                    pass
                await asyncio.sleep(0.8)
        except Exception:
            pass
        await asyncio.sleep(1.0)

    return new_count


async def phase_daangn(region: str) -> int:
    short = region.replace("경기도", "").replace("시", "").replace("군", "").strip()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
        page = await ctx.new_page()
        result = await daangn_worker(page, WEB_CATS, short)
        await browser.close()

    return result


# ─────────────────────────────────────────────
# Phase 3: DuckDuckGo HTML + 네이버 웹검색 (병렬)
# ─────────────────────────────────────────────
async def ddg_worker_html(page, categories: list, short_region: str) -> int:
    """html.duckduckgo.com/html/ — JS 불필요, 가벼움"""
    new_count = 0
    for cat in categories:
        q = f"{short_region} {cat} 010"
        url = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote(q)}&kl=kr-kr"
        try:
            await page.goto(url, timeout=18000)
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
            await asyncio.sleep(1.0)

            items = await page.evaluate("""() => {
                const results = []
                document.querySelectorAll('.result').forEach(el => {
                    const title = el.querySelector('.result__title')?.textContent?.trim() || ''
                    const snippet = el.querySelector('.result__snippet')?.textContent?.trim() || ''
                    results.push(title + ' ' + snippet)
                })
                return results
            }""")

            for text in items:
                for m in re.finditer(PHONE_RE, text):
                    if is_010(m.group(0)):
                        before = len(_records)
                        await add_record(m.group(0), cat, "DuckDuckGo", cat)
                        if len(_records) > before:
                            new_count += 1
        except Exception:
            pass
        await asyncio.sleep(0.8)

    return new_count


async def naver_worker_web(page, categories: list, short_region: str) -> int:
    new_count = 0
    for cat in categories:
        for suffix in ["010", "연락처"]:
            q = f"{short_region} {cat} {suffix}"
            url = f"https://search.naver.com/search.naver?query={urllib.parse.quote(q)}&where=web"
            try:
                await page.goto(url, timeout=18000)
                await page.wait_for_load_state("domcontentloaded", timeout=15000)
                await asyncio.sleep(0.5)
                text = await page.inner_text("body")
                for m in re.finditer(PHONE_RE, text):
                    if is_010(m.group(0)):
                        before = len(_records)
                        await add_record(m.group(0), cat, "네이버웹", cat)
                        if len(_records) > before:
                            new_count += 1
            except Exception:
                pass
            await asyncio.sleep(0.3)

    return new_count


async def phase_web(region: str) -> int:
    short = region.replace("경기도", "").replace("시", "").replace("군", "").strip()

    # 카테고리 반씩 분할
    mid = len(WEB_CATS) // 2
    cats_a = WEB_CATS[:mid]
    cats_b = WEB_CATS[mid:]

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)

        ctx_ddg  = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
        ctx_nv   = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
        ctx_ddg2 = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
        ctx_nv2  = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")

        page_ddg  = await ctx_ddg.new_page()
        page_nv   = await ctx_nv.new_page()
        page_ddg2 = await ctx_ddg2.new_page()
        page_nv2  = await ctx_nv2.new_page()

        results = await asyncio.gather(
            ddg_worker_html(page_ddg,  cats_a, short),
            naver_worker_web(page_nv,  cats_a, short),
            ddg_worker_html(page_ddg2, cats_b, short),
            naver_worker_web(page_nv2, cats_b, short),
        )

        await browser.close()

    return sum(results)


# ─────────────────────────────────────────────
# Phase 4: 네이버지도 검증
# ─────────────────────────────────────────────
async def verify_record(page, rec: dict, region: str):
    queries = []
    if rec.get("name") and len(rec["name"]) >= 2:
        queries.append(f"{rec['name']} {region}")
    queries.append(rec["phone"])

    for q in queries:
        try:
            url = f"https://m.search.naver.com/search.naver?query={urllib.parse.quote(q)}&where=m_local"
            await page.goto(url, timeout=18000)
            await page.wait_for_load_state("domcontentloaded", timeout=14000)
            await asyncio.sleep(0.5)
            html = await page.content()
            pid = PLACE_ID_RE.search(html)
            if not pid:
                continue
            rec["verified"]  = True
            rec["naver_url"] = f"https://map.naver.com/p/entry/place/{pid.group(1)}"
            nm = re.search(r'"placeName"\s*:\s*"([^"]{2,40})"', html)
            ad = re.search(r'"roadAddress"\s*:\s*"([^"]{5,80})"', html)
            if nm: rec["biz_name"] = nm.group(1)
            if ad: rec["address"]  = ad.group(1)
            break
        except Exception:
            continue


# ─────────────────────────────────────────────
# 엑셀 저장
# ─────────────────────────────────────────────
def save_excel(records: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "010번호 수집"

    headers = ["번호", "010번호", "업체명", "업종", "주소", "검증", "네이버플레이스", "수집출처"]
    hfill = PatternFill("solid", fgColor="1a365d")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    ws.row_dimensions[1].height = 26
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center

    green  = PatternFill("solid", fgColor="c6efce")
    yellow = PatternFill("solid", fgColor="ffeb9c")

    for idx, rec in enumerate(records, 1):
        row  = idx + 1
        fill = green if rec.get("verified") else yellow
        ws.row_dimensions[row].height = 20
        row_data = [
            idx,
            rec.get("phone", ""),
            rec.get("biz_name") or rec.get("name", ""),
            rec.get("category", ""),
            rec.get("address", ""),
            "✅" if rec.get("verified") else "⚠️",
            rec.get("naver_url", ""),
            rec.get("source", ""),
        ]
        left_cols = {2, 3, 4, 5, 7, 8}
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.alignment = left if col in left_cols else center
            if col == 7 and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    col_widths = [5, 16, 20, 12, 35, 6, 50, 12]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    Path(path).parent.mkdir(exist_ok=True)
    wb.save(path)


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
async def main():
    global _records
    _records = {}

    region = REGION
    print("=" * 62)
    print(f"  소상공인 010번호 수집기 v3")
    print(f"  지역: {region} | 카카오: {len(KAKAO_CATS)}카테고리 | 검증: {'ON' if DO_VERIFY else 'OFF'}")
    print("=" * 62)

    # ── Phase 1: 카카오맵 API ────────────────
    print(f"\n🟡 [1/4] 카카오맵 API 수집 중... ({len(KAKAO_CATS)}카테고리 × 5페이지)")
    kakao_new = await phase_kakao(region)
    print(f"  ✅ 카카오 완료: +{kakao_new}개  (누적 {len(_records)}개)")

    # ── Phase 2: 당근마켓 ────────────────────
    print(f"\n🥕 [2/4] 당근마켓 수집 중... ({len(WEB_CATS)}카테고리)")
    daangn_new = await phase_daangn(region)
    print(f"  ✅ 당근 완료: +{daangn_new}개  (누적 {len(_records)}개)")

    # ── Phase 3: DDG + 네이버 병렬 ──────────
    print(f"\n🔍 [3/4] DuckDuckGo + 네이버 병렬 수집 중... (4개 동시)")
    web_new = await phase_web(region)
    print(f"  ✅ 웹검색 완료: +{web_new}개  (누적 {len(_records)}개)")

    records = list(_records.values())
    print(f"\n  📊 총 수집: {len(records)}개 (중복 제거)")

    # ── Phase 4: 검증 ────────────────────────
    if DO_VERIFY and records:
        print(f"\n🗺️  [4/4] 네이버지도 검증 중... ({len(records)}개)")
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)
            ctx   = await browser.new_context(user_agent=MOBILE_UA, locale="ko-KR")
            vpage = await ctx.new_page()
            verified = 0
            for i, rec in enumerate(records, 1):
                await verify_record(vpage, rec, region)
                if rec["verified"]:
                    verified += 1
                if i % 50 == 0:
                    print(f"  [{i}/{len(records)}] 확인됨: {verified}개")
                await asyncio.sleep(0.7)
            await browser.close()
        print(f"  ✅ 검증 완료: {verified}개 확인 / {len(records)-verified}개 미확인")

    # ── 저장 ────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = f"output/{region}_010번호_v3_{timestamp}.xlsx"
    save_excel(records, out_path)

    src_stats = {}
    for r in records:
        s = r["source"]
        src_stats[s] = src_stats.get(s, 0) + 1

    print(f"\n{'='*62}")
    print(f"  완료!  파일: {out_path}")
    print(f"  총 010번호: {len(records)}개")
    for src, cnt in sorted(src_stats.items(), key=lambda x: -x[1]):
        print(f"    {src}: {cnt}개")
    print(f"{'='*62}")


if __name__ == "__main__":
    asyncio.run(main())
