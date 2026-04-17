"""
미확인 업체 재검증
사용법: python reverify.py db/포천_010번호_최종_20260326_171844.xlsx 포천
"""
import asyncio, sys, re, urllib.parse
from pathlib import Path
from datetime import datetime

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

MOBILE_UA   = ("Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
               "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1")
BROWSER_ARGS = ["--disable-gpu","--disable-dev-shm-usage","--no-sandbox","--mute-audio"]
PLACE_ID_RE  = re.compile(r'/place/(\d{6,})')


def load_unverified(path: str):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[6] == "⚠️":  # 검증 컬럼
            rec = dict(zip(hdrs, row))
            rec["verified"]      = False
            rec["biz_name"]      = rec.get("업체명", "")
            rec["name"]          = rec.get("업체명", "")
            rec["category"]      = rec.get("업종", "")
            rec["collect_address"] = rec.get("수집주소", "") or ""
            rec["naver_address"] = ""
            rec["naver_url"]     = ""
            rec["source"]        = rec.get("출처", "")
            records.append(rec)
    return records


async def _verify_one(page, rec: dict, region: str) -> bool:
    name = rec.get("name", "")
    q    = name if name else rec.get("010번호", "")
    try:
        await page.goto(
            f"https://m.search.naver.com/search.naver?query={urllib.parse.quote(q)}&where=m_local",
            timeout=20000
        )
        await page.wait_for_load_state("networkidle", timeout=15000)
        await asyncio.sleep(1.0)
        html = await page.content()
        pid  = PLACE_ID_RE.search(html)
        if pid:
            nm = re.search(r'"placeName"\s*:\s*"([^"]{2,40})"', html)
            ad = re.search(r'"roadAddress"\s*:\s*"([^"]{5,80})"', html)
            rec["verified"]      = True
            rec["naver_url"]     = f"https://map.naver.com/p/entry/place/{pid.group(1)}"
            rec["naver_address"] = ad.group(1) if ad else ""
            if nm: rec["biz_name"] = nm.group(1)
            return True
    except Exception:
        pass
    return False


async def verify_chunk(page, chunk, region, progress):
    n = 0
    for rec in chunk:
        ok = await _verify_one(page, rec, region)
        if ok: n += 1
        progress[0] += 1
        if progress[0] % 50 == 0:
            print(f"  [{progress[0]}/{progress[1]}] 확인됨: {progress[2]}개")
        progress[2] += ok
        await asyncio.sleep(0.6)
    return n


async def main():
    if len(sys.argv) < 3:
        print("사용법: python reverify.py <엑셀파일> <지역명>")
        return

    src, region = sys.argv[1], sys.argv[2]
    print(f"\n미확인 재검증: {src} / 지역: {region}")

    records = load_unverified(src)
    print(f"미확인 업체: {len(records)}개\n")

    NUM    = 3
    chunks = [records[i::NUM] for i in range(NUM)]
    prog   = [0, len(records), 0]  # [done, total, verified]

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True, args=BROWSER_ARGS)
        pages   = []
        for _ in range(NUM):
            ctx  = await browser.new_context(user_agent=MOBILE_UA, locale="ko-KR")
            page = await ctx.new_page()
            pages.append(page)

        results = await asyncio.gather(
            *[verify_chunk(pages[i], chunks[i], region, prog) for i in range(NUM)]
        )
        await browser.close()

    verified_count = sum(results)
    print(f"\n✅ 새로 확인됨: {verified_count}개 / 여전히 미확인: {len(records)-verified_count}개")

    # 원본 엑셀에 결과 반영 후 새 파일로 저장
    wb  = openpyxl.load_workbook(src)
    ws  = wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    phone_col    = hdrs.index("010번호") + 1
    verify_col   = hdrs.index("검증") + 1
    nurl_col     = hdrs.index("네이버플레이스") + 1
    naddr_col    = hdrs.index("네이버지도주소") + 1
    name_col     = hdrs.index("업체명") + 1

    verified_map = {r.get("010번호"): r for r in records if r.get("verified")}

    green  = PatternFill("solid", fgColor="c6efce")
    yellow = PatternFill("solid", fgColor="ffeb9c")

    for row in ws.iter_rows(min_row=2):
        phone = row[phone_col - 1].value
        if phone in verified_map:
            rec = verified_map[phone]
            for cell in row: cell.fill = green
            row[verify_col - 1].value  = "✅"
            row[nurl_col - 1].value    = rec["naver_url"]
            row[nurl_col - 1].hyperlink = rec["naver_url"]
            row[nurl_col - 1].font     = Font(color="0563C1", underline="single")
            row[naddr_col - 1].value   = rec["naver_address"]
            if rec.get("biz_name"):
                row[name_col - 1].value = rec["biz_name"]

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    out  = str(Path(src).parent / f"{region}_재검증완료_{ts}.xlsx")
    wb.save(out)
    print(f"저장: {out}")


if __name__ == "__main__":
    asyncio.run(main())
