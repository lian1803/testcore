"""
서브지역 카카오맵 테스트
"양주" 대신 "양주 회천", "양주 옥정" 등 동별로 검색
"""
import asyncio, re, json, sys, urllib.parse
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import httpx, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DESKTOP_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"

# 서브지역 ("양주" 제외 — main_final.py가 이미 커버)
SUB_AREAS = [
    "양주 회천", "양주 백석", "양주 덕정", "양주 덕계",
    "양주 옥정", "양주 장흥", "양주 광적", "양주 남면", "양주 은현",
]

KAKAO_CATS = [
    "치킨", "피자", "족발", "삼겹살", "분식", "중국집", "일식", "돈까스",
    "국밥", "냉면", "빵집", "도시락", "맛집", "식당", "떡볶이", "순대국",
    "미용실", "네일", "헬스장", "필라테스", "피부관리", "왁싱", "속눈썹",
    "카페", "커피숍", "버블티",
    "학원", "공부방", "과외", "레슨",
    "세탁소", "청소업체", "인테리어", "수리", "이사", "꽃집",
    "반려동물", "동물병원", "애견미용",
]

PHONE_RE = re.compile(r'01[016789][-.\s]?\d{3,4}[-.\s]?\d{4}')
_lock = asyncio.Lock()
_records: dict[str, dict] = {}

def normalize(raw):
    d = re.sub(r'[^\d]', '', raw)
    if len(d) == 11: return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if len(d) == 10: return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    return raw

def is_010(raw):
    d = re.sub(r'[^\d]', '', raw)
    return d.startswith('010') and len(d) == 11

async def add(phone_raw, name, area, cat):
    if not is_010(phone_raw): return
    phone = normalize(phone_raw)
    async with _lock:
        if phone not in _records:
            _records[phone] = {"phone": phone, "name": name, "area": area, "category": cat, "source": "카카오맵(서브지역)"}

async def kakao_one(client, sem, query, page, area, cat):
    url = (f"https://search.map.kakao.com/mapsearch/map.daum"
           f"?callback=cb&q={urllib.parse.quote(query)}&page={page}&size=15&sort=0")
    hdrs = {"User-Agent": DESKTOP_UA, "Referer": "https://map.kakao.com/", "Accept-Language": "ko-KR,ko;q=0.9"}
    async with sem:
        try:
            r = await client.get(url, headers=hdrs, timeout=10)
            js = re.sub(r'^/\*\*/cb\(', '', r.text.strip())
            js = re.sub(r'\);\s*$', '', js)
            places = json.loads(js).get("place", [])
        except:
            return 0
        await asyncio.sleep(0.15)
    cnt = 0
    for p in places:
        tel = p.get("tel", "")
        if is_010(tel):
            before = len(_records)
            await add(tel, p.get("name", ""), area, cat)
            if len(_records) > before: cnt += 1
    return cnt

async def main():
    global _records
    _records = {}
    total_queries = len(SUB_AREAS) * len(KAKAO_CATS) * 10
    print("=" * 58)
    print(f"  서브지역 카카오맵 테스트")
    print(f"  서브지역: {len(SUB_AREAS)}개 | 카테고리: {len(KAKAO_CATS)}개 | 총 쿼리: {total_queries}개")
    print("=" * 58)

    sem = asyncio.Semaphore(10)
    tasks = []
    async with httpx.AsyncClient() as client:
        for area in SUB_AREAS:
            for cat in KAKAO_CATS:
                q = f"{area} {cat}"
                for pg in range(1, 11):
                    tasks.append(kakao_one(client, sem, q, pg, area, cat))
        results = await asyncio.gather(*tasks)

    total = sum(results)
    records = list(_records.values())
    print(f"\n  완료: {len(records)}개 (서브지역 신규, 중복제거)")

    # 엑셀 저장
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "서브지역 카카오"
    hdrs = ["번호", "010번호", "업체명", "업종", "지역"]
    hfill = PatternFill("solid", fgColor="1a365d")
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    yellow = PatternFill("solid", fgColor="ffeb9c")
    for idx, rec in enumerate(records, 1):
        ws.row_dimensions[idx+1].height = 18
        for col, val in enumerate([idx, rec["phone"], rec["name"], rec["category"], rec["area"]], 1):
            cell = ws.cell(row=idx+1, column=col, value=val)
            cell.fill = yellow

    for col, w in enumerate([5, 16, 20, 12, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = f"output/양주_서브지역테스트_{ts}.xlsx"
    Path("output").mkdir(exist_ok=True)
    wb.save(out)
    print(f"  파일: {out}")
    print("=" * 58)

if __name__ == "__main__":
    asyncio.run(main())
