"""
소상공인 010번호 수집기 v2
===========================
v1 대비 변경:
  - 병렬 처리: 브라우저 5개 동시 실행 (~5배 속도)
  - DuckDuckGo 추가: API방식, Playwright 불필요, 구글급 인덱싱
  - 쿼리 다양화: "010" / "연락처" / "인스타" 등 4가지 패턴
  - 결과: 50분에 2000~4000개 목표

사용법:
  python main_v2.py              ← 양주, 검증 없음
  python main_v2.py 의정부
  python main_v2.py 포천 --verify  ← 검증 포함 (느림)
"""

import asyncio
import re
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, TimeoutError as PWTimeout
from ddgs import DDGS

# ─────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────
REGION   = sys.argv[1] if len(sys.argv) > 1 else "양주"
DO_VERIFY = "--verify" in sys.argv

NAVER_CLIENT_ID     = "o0776HJDmQVO6J9Lez1m"
NAVER_CLIENT_SECRET = "ZXG0lPbgH9"

DESKTOP_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
MOBILE_UA = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
)

PARALLEL_WORKERS = 5   # 동시 브라우저 수
MAX_PAGES_WEB    = 5   # 페이지당 검색 페이지 수 (v1=3)

SUB_AREAS = {
    "양주": [
        "양주", "양주 회천", "양주 백석", "양주 덕정", "양주 덕계",
        "양주 옥정", "양주 장흥", "양주 광적", "양주 남면", "양주 은현",
    ],
    "의정부": [
        "의정부", "의정부 가능", "의정부 녹양", "의정부 신곡",
        "의정부 호원", "의정부 장암", "의정부 민락", "의정부 흥선",
    ],
    "포천": [
        "포천", "포천 소흘", "포천 군내", "포천 동면",
        "포천 신북", "포천 가산", "포천 일동", "포천 영중",
    ],
}

CATEGORIES = [
    "음식점", "카페", "미용실", "헤어샵", "네일샵", "네일아트",
    "헬스장", "필라테스", "요가", "피티",
    "병원", "치과", "한의원", "피부과", "안과", "정형외과",
    "학원", "태권도", "피아노", "영어학원",
    "꽃집", "베이커리", "빵집", "케이크",
    "부동산", "공인중개사",
    "펜션", "게스트하우스",
    "세탁소", "사진관", "애견미용", "반려동물",
    "인테리어", "리모델링",
    "치킨", "피자", "족발", "보쌈",
]

# 쿼리 패턴 (다양화로 더 많은 페이지 커버)
QUERY_PATTERNS = [
    "{area} {cat} 010",
    "{area} {cat} 연락처",
    "{area} {cat} 인스타",
    "{area} {cat} 전화번호",
]

FRANCHISE_EXCLUDE = {
    "스타벅스", "이마트", "홈플러스", "롯데마트", "코스트코", "다이소",
    "맥도날드", "버거킹", "롯데리아", "KFC", "서브웨이", "맘스터치",
    "GS25", "CU", "세븐일레븐", "미니스톱", "이마트24",
    "올리브영", "유니클로", "자라", "H&M", "크린토피아",
    "bbq", "굽네치킨", "교촌치킨", "bhc",
}

PHONE_RE   = re.compile(r'01[016789][-.\s]?\d{3,4}[-.\s]?\d{4}')
PLACE_ID_RE = re.compile(r'/place/(\d{6,})')

# 공유 수집 결과 (thread-safe via asyncio lock)
_lock   = asyncio.Lock()
_records: dict[str, dict] = {}

# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────
def normalize_phone(raw: str) -> str:
    d = re.sub(r'[^\d]', '', raw)
    if len(d) == 11: return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if len(d) == 10: return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    return raw

def is_010(raw: str) -> bool:
    return re.sub(r'[^\d]', '', raw).startswith('010')

def extract_name_hint(text: str, m) -> str:
    before = text[max(0, m.start() - 60):m.start()].strip()
    words  = re.findall(r'[가-힣A-Za-z0-9]{2,20}', before)
    return words[-1] if words else ""

async def add_record(phone_raw: str, name: str, source: str, query: str):
    phone = normalize_phone(phone_raw)
    if not is_010(phone_raw):
        return
    async with _lock:
        if phone not in _records:
            _records[phone] = {
                "phone": phone, "name": name,
                "source": source, "query": query,
                "verified": False, "biz_name": "",
                "address": "", "naver_url": "",
            }

# ─────────────────────────────────────────────
# Phase 1: DuckDuckGo (빠른 API, 병렬 가능)
# ─────────────────────────────────────────────
def _ddg_search(q: str) -> list:
    """동기 DDG 검색 — run_in_executor로 호출해야 함"""
    try:
        return list(DDGS().text(q, region="kr-kr", max_results=30))
    except Exception:
        return []


async def ddg_worker(queries: list[str], worker_id: int, semaphore: asyncio.Semaphore):
    loop = asyncio.get_event_loop()
    new_count = 0
    for q in queries:
        async with semaphore:
            try:
                results = await loop.run_in_executor(None, _ddg_search, q)
                for r in results:
                    text = r.get("title", "") + " " + r.get("body", "")
                    for m in PHONE_RE.finditer(text):
                        if is_010(m.group(0)):
                            before_count = len(_records)
                            await add_record(m.group(0), extract_name_hint(text, m), "DuckDuckGo", q)
                            if len(_records) > before_count:
                                new_count += 1
            except Exception:
                pass
            await asyncio.sleep(0.05)
    return new_count


async def phase_ddg(queries: list[str]) -> int:
    semaphore = asyncio.Semaphore(8)  # DDG는 API라서 병렬 더 많이 허용
    chunk = max(1, len(queries) // 8)
    chunks = [queries[i:i+chunk] for i in range(0, len(queries), chunk)]

    tasks = [ddg_worker(c, i, semaphore) for i, c in enumerate(chunks)]
    results = await asyncio.gather(*tasks)
    return sum(results)


# ─────────────────────────────────────────────
# Phase 2: 네이버 웹검색 (병렬 Playwright)
# ─────────────────────────────────────────────
async def naver_worker(page, queries: list[str], source: str = "네이버"):
    new_count = 0
    for query in queries:
        for pg in range(MAX_PAGES_WEB):
            start = pg * 10 + 1
            encoded = urllib.parse.quote(query)
            try:
                await page.goto(
                    f"https://search.naver.com/search.naver?query={encoded}&where=web&start={start}",
                    timeout=18000
                )
                await page.wait_for_load_state("domcontentloaded", timeout=14000)
                await page.wait_for_timeout(600)
                text = await page.inner_text("body")

                if "검색결과가 없습니다" in text:
                    break

                for m in PHONE_RE.finditer(text):
                    if is_010(m.group(0)):
                        before = len(_records)
                        await add_record(m.group(0), extract_name_hint(text, m), source, query)
                        if len(_records) > before:
                            new_count += 1
                await asyncio.sleep(0.25)
            except Exception:
                break
    return new_count


async def daum_worker(page, queries: list[str]):
    new_count = 0
    for query in queries:
        for pg in range(MAX_PAGES_WEB):
            encoded = urllib.parse.quote(query)
            try:
                await page.goto(
                    f"https://search.daum.net/search?w=web&q={encoded}&p={pg+1}",
                    timeout=18000
                )
                await page.wait_for_load_state("domcontentloaded", timeout=14000)
                await page.wait_for_timeout(600)
                text = await page.inner_text("body")

                if "검색결과가 없습니다" in text:
                    break

                for m in PHONE_RE.finditer(text):
                    if is_010(m.group(0)):
                        before = len(_records)
                        await add_record(m.group(0), extract_name_hint(text, m), "다음(카카오)", query)
                        if len(_records) > before:
                            new_count += 1
                await asyncio.sleep(0.25)
            except Exception:
                break
    return new_count


async def phase_browser(all_queries: list[str]) -> int:
    """네이버 + 다음을 PARALLEL_WORKERS개 브라우저로 병렬 처리"""
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)

        # 네이버용 컨텍스트 (PARALLEL_WORKERS개)
        naver_ctxs  = [
            await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
            for _ in range(PARALLEL_WORKERS)
        ]
        naver_pages = [await ctx.new_page() for ctx in naver_ctxs]

        # 다음용 컨텍스트 (PARALLEL_WORKERS개)
        daum_ctxs  = [
            await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
            for _ in range(PARALLEL_WORKERS)
        ]
        daum_pages = [await ctx.new_page() for ctx in daum_ctxs]

        # 쿼리를 워커수로 균등 분할
        chunk = max(1, len(all_queries) // PARALLEL_WORKERS)
        chunks = [all_queries[i:i+chunk] for i in range(0, len(all_queries), chunk)]

        naver_tasks = [naver_worker(naver_pages[i], chunks[i]) for i in range(min(PARALLEL_WORKERS, len(chunks)))]
        daum_tasks  = [daum_worker(daum_pages[i],  chunks[i]) for i in range(min(PARALLEL_WORKERS, len(chunks)))]

        results = await asyncio.gather(*(naver_tasks + daum_tasks))

        await browser.close()
    return sum(results)


# ─────────────────────────────────────────────
# Phase 3: 네이버 블로그 API (빠름)
# ─────────────────────────────────────────────
async def phase_blog_api(areas: list[str]) -> int:
    headers = {
        "X-Naver-Client-Id":     NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    new_count = 0
    async with httpx.AsyncClient(timeout=10) as client:
        for area in areas[:4]:
            for cat in CATEGORIES:
                for suffix in ["010", "연락처", "전화번호"]:
                    query = f"{area} {cat} {suffix}"
                    for start in range(1, 51, 10):
                        try:
                            resp = await client.get(
                                "https://openapi.naver.com/v1/search/blog.json",
                                headers=headers,
                                params={"query": query, "display": 10, "start": start, "sort": "date"},
                            )
                            items = resp.json().get("items", [])
                            if not items:
                                break
                            for item in items:
                                text = re.sub(r'<[^>]+>', '',
                                    item.get("title", "") + " " + item.get("description", ""))
                                for m in PHONE_RE.finditer(text):
                                    if is_010(m.group(0)):
                                        before = len(_records)
                                        await add_record(m.group(0), extract_name_hint(text, m), "네이버블로그", query)
                                        if len(_records) > before:
                                            new_count += 1
                            await asyncio.sleep(0.1)
                        except Exception:
                            break
    return new_count


# ─────────────────────────────────────────────
# Phase 4: 네이버지도 검증
# ─────────────────────────────────────────────
async def verify_record(page, rec: dict, region: str):
    name_hint = rec.get("name", "")
    phone     = rec["phone"]
    queries   = []
    if name_hint and len(name_hint) >= 2:
        queries.append(f"{name_hint} {region}")
    queries.append(phone)

    for q in queries:
        try:
            encoded = urllib.parse.quote(q)
            await page.goto(
                f"https://m.search.naver.com/search.naver?query={encoded}&where=m_local",
                timeout=18000
            )
            await page.wait_for_load_state("domcontentloaded", timeout=14000)
            await page.wait_for_timeout(500)
            html = await page.content()
            text = await page.inner_text("body")
            pid  = PLACE_ID_RE.search(html)
            if not pid:
                continue
            rec["verified"]   = True
            rec["naver_url"]  = f"https://map.naver.com/p/entry/place/{pid.group(1)}"
            nm = re.search(r'"placeName"\s*:\s*"([^"]{2,40})"', html)
            ad = re.search(r'"roadAddress"\s*:\s*"([^"]{5,80})"', html)
            if nm: rec["biz_name"] = nm.group(1)
            if ad: rec["address"]  = ad.group(1)
            break
        except Exception:
            continue


# ─────────────────────────────────────────────
# 엑셀 저장
# ─────────────────────────────────────────────
def save_excel(records: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "010번호 수집"

    headers = ["번호", "010번호", "업체명", "주소", "검증여부", "네이버플레이스", "수집출처", "쿼리"]
    hfill = PatternFill("solid", fgColor="1a365d")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    ws.row_dimensions[1].height = 26
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center

    green  = PatternFill("solid", fgColor="c6efce")
    yellow = PatternFill("solid", fgColor="ffeb9c")

    for idx, rec in enumerate(records, 1):
        row  = idx + 1
        fill = green if rec.get("verified") else yellow
        ws.row_dimensions[row].height = 20
        row_data = [
            idx,
            rec.get("phone", ""),
            rec.get("biz_name") or rec.get("name", ""),
            rec.get("address", ""),
            "✅ 확인" if rec.get("verified") else "⚠️ 미확인",
            rec.get("naver_url", ""),
            rec.get("source", ""),
            rec.get("query", ""),
        ]
        left_cols = {2, 3, 4, 6, 7, 8}
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.alignment = left if col in left_cols else center
            if col == 6 and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    col_widths = [5, 16, 20, 35, 10, 50, 14, 35]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    Path(path).parent.mkdir(exist_ok=True)
    wb.save(path)


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
async def main():
    global _records
    _records = {}

    region = REGION
    areas  = SUB_AREAS.get(region, [region])

    # 전체 쿼리 생성
    all_queries = []
    for area in areas:
        for cat in CATEGORIES:
            for pat in QUERY_PATTERNS:
                all_queries.append(pat.format(area=area, cat=cat))

    print("=" * 62)
    print(f"  소상공인 010번호 수집기 v2")
    print(f"  지역: {region} | 서브지역: {len(areas)}개 | 업종: {len(CATEGORIES)}개")
    print(f"  총 쿼리: {len(all_queries)}개 | 병렬: {PARALLEL_WORKERS}개")
    print(f"  검증: {'ON' if DO_VERIFY else 'OFF (--verify로 켜기)'}")
    print("=" * 62)

    # ── Phase 1: DuckDuckGo ──────────────────
    print(f"\n🦆 [1/4] DuckDuckGo 수집 중... ({len(all_queries)}개 쿼리)\n")
    ddg_new = await phase_ddg(all_queries)
    print(f"  ✅ DDG 완료: +{ddg_new}개  (누적 {len(_records)}개)")

    # ── Phase 2: 네이버 + 다음 병렬 ─────────
    # 쿼리 중 "010" 패턴만 선별 (다른 건 DDG가 이미 커버)
    web_queries = [q for q in all_queries if "010" in q or "연락처" in q]
    print(f"\n🔍 [2/4] 네이버+다음 병렬 수집... ({len(web_queries)}개 × {PARALLEL_WORKERS}병렬)\n")
    web_new = await phase_browser(web_queries)
    print(f"  ✅ 웹검색 완료: +{web_new}개  (누적 {len(_records)}개)")

    # ── Phase 3: 블로그 API ──────────────────
    print(f"\n📝 [3/4] 네이버 블로그 API 수집 중...\n")
    blog_new = await phase_blog_api(areas)
    print(f"  ✅ 블로그 완료: +{blog_new}개  (누적 {len(_records)}개)")

    records = list(_records.values())
    print(f"\n  📊 총 수집: {len(records)}개 (010번호 기준 중복 제거)")

    # ── Phase 4: 검증 (옵션) ─────────────────
    if DO_VERIFY and records:
        print(f"\n🗺️  [4/4] 네이버지도 검증 중... ({len(records)}개)\n")
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)
            ctx   = await browser.new_context(user_agent=MOBILE_UA, locale="ko-KR")
            vpage = await ctx.new_page()
            verified = 0
            for i, rec in enumerate(records, 1):
                await verify_record(vpage, rec, region)
                if rec["verified"]:
                    verified += 1
                if i % 50 == 0:
                    print(f"  [{i}/{len(records)}] 확인됨: {verified}개")
                await asyncio.sleep(0.7)
            await browser.close()
        print(f"  ✅ 검증 완료: {verified}개 확인 / {len(records)-verified}개 미확인")

    # ── 저장 ────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = f"output/{region}_010번호_v2_{timestamp}.xlsx"
    print(f"\n💾 엑셀 저장 중...")
    save_excel(records, out_path)

    verified_n = sum(1 for r in records if r.get("verified"))
    print(f"\n{'='*62}")
    print(f"  완료! 파일: {out_path}")
    print(f"  총 010번호: {len(records)}개")
    if DO_VERIFY:
        print(f"  네이버지도 확인: {verified_n}개")
    print(f"{'='*62}")


if __name__ == "__main__":
    asyncio.run(main())
