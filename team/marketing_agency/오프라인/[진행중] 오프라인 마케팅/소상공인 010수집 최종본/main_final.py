"""
소상공인 010번호 수집기 — 최종본
=================================
소스 (전부 동시 실행):
  1. 카카오맵 API   — JSONP, asyncio 10개 동시, Playwright 없음 (가장 많이 나옴)
  2. 네이버맵 API   — map.naver.com allSearch 인터셉트, 구조화 데이터
  3. 당근마켓       — /local-profile/ 3개 브라우저 병렬
  4. 네이버 웹검색  — 4개 브라우저 병렬
  5. DuckDuckGo     — 2개 브라우저 병렬

제거: 네이버블로그 (6개밖에 안 나옴), 다음검색 (네이버와 중복)

사용법:
  python main_final.py            ← 양주 (검증 포함)
  python main_final.py 의정부
  python main_final.py 포천 --no-verify   ← 검증 스킵
"""

import asyncio
import json
import re
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ─────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────
REGION    = next((a for a in sys.argv[1:] if not a.startswith("--")), "양주")
DO_VERIFY = "--no-verify" not in sys.argv   # 기본 ON, --no-verify 로 스킵
ONLY      = next((a.split("=")[1] for a in sys.argv if a.startswith("--only=")), None)

# 지역별 서브지역 (카카오맵 검색에 전부 사용)
SUB_AREAS = {
    "양주": [
        "양주", "양주 회천", "양주 백석", "양주 덕정", "양주 덕계",
        "양주 옥정", "양주 장흥", "양주 광적", "양주 남면", "양주 은현",
    ],
    "의정부": [
        "의정부", "의정부 가능", "의정부 녹양", "의정부 신곡",
        "의정부 호원", "의정부 장암", "의정부 민락", "의정부 흥선",
    ],
    "포천": [
        "포천", "포천 소흘", "포천 군내", "포천 동면",
        "포천 신북", "포천 가산", "포천 일동", "포천 영중",
    ],
    "남양주": [
        "남양주", "남양주 화도", "남양주 수동", "남양주 조안",
        "남양주 와부", "남양주 진접", "남양주 별내", "남양주 퇴계원",
        "남양주 진건", "남양주 오남",
    ],
    "가평": [
        "가평", "가평 청평", "가평 설악", "가평 조종",
        "가평 북면", "가평 상면",
    ],
    "연천": [
        "연천", "연천 전곡", "연천 청산", "연천 백학",
        "연천 미산", "연천 왕징",
    ],
    "동두천": [
        "동두천", "동두천 생연", "동두천 보산", "동두천 송내",
        "동두천 지행", "동두천 탑동",
    ],
    "구리": [
        "구리", "구리 인창", "구리 교문", "구리 수택",
        "구리 갈매",
    ],
    "하남": [
        "하남", "하남 미사", "하남 풍산", "하남 덕풍",
        "하남 신장", "하남 망월",
    ],
    "양평": [
        "양평", "양평 양서", "양평 옥천", "양평 강상",
        "양평 강하", "양평 청운", "양평 단월",
    ],
    "파주": [
        "파주", "파주 금촌", "파주 문산", "파주 운정",
        "파주 교하", "파주 적성", "파주 법원", "파주 탄현",
    ],
    "고양": [
        "고양", "고양 일산", "고양 화정", "고양 능곡",
        "고양 원당", "고양 행신", "고양 탄현", "고양 식사",
    ],
    "성남": [
        "성남", "성남 분당", "성남 수정", "성남 중원",
        "성남 판교", "성남 야탑", "성남 모란", "성남 서현",
    ],
    "광주": [
        "광주", "광주 오포", "광주 초월", "광주 곤지암",
        "광주 실촌", "광주 남종",
    ],
    "이천": [
        "이천", "이천 부발", "이천 장호원", "이천 율면",
        "이천 백사", "이천 설성",
    ],
    "여주": [
        "여주", "여주 가남", "여주 점동", "여주 흥천",
        "여주 금사", "여주 능서",
    ],
    # ── 경기 남부/서부 ──────────────────────────
    "수원": [
        "수원", "수원 장안", "수원 권선", "수원 팔달", "수원 영통",
        "수원 광교", "수원 인계", "수원 매탄",
    ],
    "용인": [
        "용인", "용인 수지", "용인 기흥", "용인 처인",
        "용인 동백", "용인 죽전", "용인 보라", "용인 역북",
    ],
    "안산": [
        "안산", "안산 단원", "안산 상록", "안산 고잔",
        "안산 본오", "안산 중앙", "안산 초지",
    ],
    "안양": [
        "안양", "안양 만안", "안양 동안", "안양 평촌",
        "안양 비산", "안양 호계",
    ],
    "부천": [
        "부천", "부천 원미", "부천 소사", "부천 오정",
        "부천 중동", "부천 상동", "부천 범박",
    ],
    "시흥": [
        "시흥", "시흥 정왕", "시흥 배곧", "시흥 목감",
        "시흥 은계", "시흥 장현",
    ],
    "화성": [
        "화성", "화성 동탄", "화성 병점", "화성 봉담",
        "화성 향남", "화성 남양", "화성 태안", "화성 우정",
    ],
    "평택": [
        "평택", "평택 고덕", "평택 서정", "평택 송탄",
        "평택 안중", "평택 통복", "평택 비전",
    ],
    "김포": [
        "김포", "김포 구래", "김포 장기", "김포 풍무",
        "김포 통진", "김포 양촌", "김포 고촌",
    ],
    "광명": [
        "광명", "광명 철산", "광명 하안", "광명 소하",
    ],
    "군포": [
        "군포", "군포 산본", "군포 당동", "군포 둔대",
    ],
    "의왕": [
        "의왕", "의왕 내손", "의왕 오전",
    ],
    "오산": [
        "오산", "오산 세마", "오산 초평", "오산 대원",
    ],
    "과천": [
        "과천", "과천 별양", "과천 문원",
    ],
    "안성": [
        "안성", "안성 공도", "안성 보개", "안성 미양", "안성 죽산",
    ],
    # ── 충청 ────────────────────────────────────
    "천안": [
        "천안", "천안 쌍용", "천안 불당", "천안 두정",
        "천안 신부", "천안 성정", "천안 백석", "천안 직산",
    ],
    "아산": [
        "아산", "아산 배방", "아산 탕정", "아산 온양",
        "아산 신창", "아산 염치",
    ],
    "청주": [
        "청주", "청주 흥덕", "청주 상당", "청주 청원",
        "청주 서원", "청주 율량", "청주 오창", "청주 오송",
    ],
    "충주": [
        "충주", "충주 중앙", "충주 연수", "충주 용산",
        "충주 호암", "충주 주덕",
    ],
    "공주": [
        "공주", "공주 신관", "공주 웅진",
    ],
    "보령": [
        "보령", "보령 대천", "보령 주교", "보령 웅천",
    ],
    "서산": [
        "서산", "서산 동문", "서산 석남", "서산 대산", "서산 해미",
    ],
    "당진": [
        "당진", "당진 합덕", "당진 면천", "당진 신평",
    ],
    "홍성": [
        "홍성", "홍성 홍북", "홍성 광천",
    ],
    "세종": [
        "세종", "세종 조치원", "세종 고운", "세종 한솔",
        "세종 도담", "세종 아름", "세종 종촌",
    ],
    # ── 대전 ────────────────────────────────────
    "대전": [
        "대전", "대전 유성", "대전 서구", "대전 중구",
        "대전 동구", "대전 대덕", "대전 둔산", "대전 노은",
        "대전 관평", "대전 봉명",
    ],
}

DESKTOP_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)
MOBILE_UA = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
)

# ── 카카오맵 카테고리 (39개) ──────────────────
KAKAO_CATS = [
    # 음식
    "치킨", "피자", "족발", "삼겹살", "분식", "중국집", "일식", "돈까스",
    "국밥", "냉면", "빵집", "도시락", "맛집", "식당", "떡볶이", "순대국",
    "곱창", "막창", "칼국수", "해장국", "초밥", "라멘",
    "마라탕", "양꼬치", "쌀국수", "베트남음식",
    "술집", "포차", "막걸리",
    # 뷰티/헬스
    "미용실", "네일", "헬스장", "필라테스", "피부관리", "왁싱", "속눈썹",
    "타투", "반영구화장", "눈썹문신",
    "마사지", "스포츠마사지",
    "요가", "크로스핏", "PT",
    # 카페
    "카페", "커피숍", "버블티",
    # 교육
    "학원", "공부방", "과외", "레슨",
    "피아노", "기타레슨", "음악학원",
    "태권도", "검도", "무술",
    "미술학원", "영어학원", "수학학원",
    # 서비스
    "세탁소", "청소업체", "인테리어", "수리", "이사", "꽃집",
    "자동차정비", "세차", "핸드폰수리", "컴퓨터수리",
    "에어컨청소", "보일러수리", "열쇠",
    "부동산", "사진관", "대리운전",
    # 반려동물
    "반려동물", "동물병원", "애견미용",
]

# ── 네이버맵 / 당근 / 웹 검색 카테고리 ────────
WEB_CATS = [
    "음식점", "카페", "미용실", "네일", "헬스장", "필라테스",
    "학원", "병원", "치과", "한의원", "피부과",
    "세탁소", "꽃집", "베이커리", "빵집",
    "치킨", "피자", "족발", "보쌈", "분식",
    "곱창", "마라탕", "칼국수", "술집",
    "타투", "마사지", "요가", "크로스핏",
    "피아노", "태권도", "미술학원",
    "인테리어", "반려동물", "애견미용",
    "부동산", "사진관", "대리운전",
    "자동차정비", "핸드폰수리", "컴퓨터수리",
]

NAVER_MAP_CATS = [
    "카페", "음식점", "미용실", "네일샵", "헬스장", "필라테스",
    "학원", "병원", "치과", "세탁소", "꽃집",
    "치킨", "피자", "족발", "분식",
    "인테리어", "반려동물",
]

FRANCHISE_EXCLUDE = {
    "스타벅스", "이마트", "홈플러스", "롯데마트", "코스트코", "다이소",
    "맥도날드", "버거킹", "롯데리아", "KFC", "서브웨이", "맘스터치",
    "GS25", "CU", "세븐일레븐", "미니스톱", "이마트24",
    "올리브영", "유니클로", "자라", "H&M", "크린토피아",
    "bbq", "굽네치킨", "교촌치킨", "bhc",
}

PHONE_RE    = re.compile(r'01[016789][-.\s]?\d{3,4}[-.\s]?\d{4}')
PLACE_ID_RE = re.compile(r'/place/(\d{6,})')

# CPU 절약용 브라우저 실행 인자
BROWSER_ARGS = [
    "--disable-gpu", "--disable-software-rasterizer",
    "--disable-dev-shm-usage", "--no-sandbox",
    "--disable-extensions", "--disable-background-networking",
    "--disable-default-apps", "--mute-audio",
]
# 차단할 리소스 타입 (이미지·폰트·스타일시트·미디어)
_BLOCK_TYPES = {"image", "font", "stylesheet", "media"}

async def block_resources(page):
    """브라우저 페이지에서 이미지/폰트/CSS 차단 → CPU/메모리 절약"""
    await page.route("**/*", lambda route: (
        route.abort() if route.request.resource_type in _BLOCK_TYPES else route.continue_()
    ))

_lock    = asyncio.Lock()
_records: dict[str, dict] = {}

# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────
def normalize(raw: str) -> str:
    d = re.sub(r'[^\d]', '', raw)
    if len(d) == 11: return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if len(d) == 10: return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    return raw

def is_010(raw: str) -> bool:
    d = re.sub(r'[^\d]', '', raw)
    return d.startswith('010') and len(d) == 11

def is_franchise(name: str) -> bool:
    return any(f in name for f in FRANCHISE_EXCLUDE)

async def add(phone_raw: str, name: str, source: str, cat: str = "", collect_address: str = ""):
    if not is_010(phone_raw): return
    phone = normalize(phone_raw)
    if is_franchise(name): return
    async with _lock:
        if phone not in _records:
            _records[phone] = {
                "phone": phone, "name": name, "category": cat,
                "source": source, "verified": False,
                "biz_name": "", "collect_address": collect_address,
                "naver_address": "", "naver_url": "",
            }


# ─────────────────────────────────────────────
# 1. 카카오맵 API (HTTP 10개 병렬)
# ─────────────────────────────────────────────
async def _kakao_one(client: httpx.AsyncClient, sem: asyncio.Semaphore,
                     query: str, page: int, cat: str) -> int:
    url = (
        "https://search.map.kakao.com/mapsearch/map.daum"
        f"?callback=cb&q={urllib.parse.quote(query)}&page={page}&size=15&sort=0"
    )
    hdrs = {"User-Agent": DESKTOP_UA, "Referer": "https://map.kakao.com/",
            "Accept-Language": "ko-KR,ko;q=0.9"}
    async with sem:
        try:
            resp = await client.get(url, headers=hdrs, timeout=10)
            js = re.sub(r'^/\*\*/cb\(', '', resp.text.strip())
            js = re.sub(r'\);\s*$', '', js)
            places = json.loads(js).get("place", [])
        except Exception:
            return 0
        await asyncio.sleep(0.2)

    cnt = 0
    for p in places:
        tel = p.get("tel", "")
        if is_010(tel):
            addr = p.get("road_address") or p.get("address", "")
            before = len(_records)
            await add(tel, p.get("name", ""), "카카오맵", cat, addr)
            if len(_records) > before: cnt += 1
    return cnt


async def scrape_kakao(region: str) -> int:
    areas = SUB_AREAS.get(region, [region])
    sem = asyncio.Semaphore(10)
    tasks = []
    async with httpx.AsyncClient() as client:
        for area in areas:
            for cat in KAKAO_CATS:
                q = f"{area} {cat}"
                for pg in range(1, 16):  # 15페이지 × 15개 = 카테고리당 최대 225개
                    tasks.append(_kakao_one(client, sem, q, pg, cat))
        results = await asyncio.gather(*tasks)
    return sum(results)


# ─────────────────────────────────────────────
# 2. 네이버맵 API 인터셉트 (allSearch 응답 가로채기)
# ─────────────────────────────────────────────
async def _nmap_search(page, query: str) -> int:
    pending_responses = []

    # 동기 핸들러로 response 객체만 저장 (await 없음)
    def on_response(resp):
        if "allSearch" in resp.url:
            pending_responses.append(resp)

    page.on("response", on_response)
    try:
        await page.goto(
            f"https://map.naver.com/p/search/{urllib.parse.quote(query)}",
            timeout=20000
        )
        await page.wait_for_load_state("domcontentloaded", timeout=15000)
        await asyncio.sleep(2.0)
        # 스크롤로 추가 로드
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await asyncio.sleep(1.5)
    except Exception:
        pass
    finally:
        page.remove_listener("response", on_response)

    cnt = 0
    for resp in pending_responses:
        try:
            data = await resp.json()
            items = (data.get("result") or {}).get("place", {})
            if isinstance(items, dict):
                items = items.get("list", [])
            elif not isinstance(items, list):
                items = []
            for item in items:
                tel = item.get("tel", "")
                if is_010(tel):
                    addr = item.get("roadAddress") or item.get("address", "")
                    before = len(_records)
                    await add(tel, item.get("name", ""), "네이버맵", "", addr)
                    if len(_records) > before: cnt += 1
        except Exception:
            pass
    return cnt


async def scrape_naver_map(short: str) -> int:
    total = 0
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True, args=BROWSER_ARGS)
        ctx  = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR",
                                          viewport={"width": 1280, "height": 800})
        ctx2 = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR",
                                          viewport={"width": 1280, "height": 800})
        page1 = await ctx.new_page()
        page2 = await ctx2.new_page()
        await block_resources(page1)
        await block_resources(page2)

        mid = len(NAVER_MAP_CATS) // 2
        cats_a = NAVER_MAP_CATS[:mid]
        cats_b = NAVER_MAP_CATS[mid:]

        async def worker(page, cats):
            n = 0
            for cat in cats:
                n += await _nmap_search(page, f"{short} {cat}")
                await asyncio.sleep(0.5)
            return n

        r1, r2 = await asyncio.gather(worker(page1, cats_a), worker(page2, cats_b))
        total = r1 + r2
        await browser.close()
    return total


# ─────────────────────────────────────────────
# 3. 당근마켓 (3개 브라우저 병렬)
# ─────────────────────────────────────────────
async def _daangn_worker(page, cats: list, short: str) -> int:
    cnt = 0
    for cat in cats:
        q = f"{short} {cat}"
        url = f"https://www.daangn.com/kr/local-profile/s/?search={urllib.parse.quote(q)}"
        try:
            await page.goto(url, timeout=15000)
            await page.wait_for_load_state("domcontentloaded", timeout=12000)
            await asyncio.sleep(1.0)

            links = await page.evaluate("""() =>
                Array.from(document.querySelectorAll('a[href*="/local-profile/"]'))
                    .map(a => a.href)
                    .filter(h => !/\\/local-profile\\/s\\//.test(h))
                    .filter((v,i,a) => a.indexOf(v)===i)
                    .slice(0,5)
            """)

            for link in links:
                try:
                    await page.goto(link, timeout=12000)
                    await page.wait_for_load_state("domcontentloaded", timeout=10000)
                    await asyncio.sleep(0.6)
                    data = await page.evaluate("""() => {
                        const h = document.querySelector('h1,h2')?.textContent?.trim()||''
                        const body = document.body.innerText
                        const phones = body.match(/010[-\\s]?\\d{3,4}[-\\s]?\\d{4}/g)||[]
                        const tels = Array.from(document.querySelectorAll('a[href^="tel:"]'))
                            .map(a=>a.href.replace('tel:','').trim())
                        return {name:h, phones, tels}
                    }""")
                    for ph in list(set(data["phones"] + [t for t in data["tels"] if is_010(t)])):
                        before = len(_records)
                        await add(ph, data["name"], "당근마켓", cat)
                        if len(_records) > before: cnt += 1
                except Exception:
                    pass
                await asyncio.sleep(0.3)
        except Exception:
            pass
        await asyncio.sleep(0.5)
    return cnt


async def scrape_daangn(short: str) -> int:
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True, args=BROWSER_ARGS)
        ctx  = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
        page = await ctx.new_page()
        await block_resources(page)
        try:
            r = await asyncio.wait_for(_daangn_worker(page, WEB_CATS, short), timeout=480)  # 8분 max
        except asyncio.TimeoutError:
            print("  ⚠️ 당근마켓 시간초과 (8분)")
            r = 0
        await browser.close()
    return r


# ─────────────────────────────────────────────
# 4. 네이버 웹검색 (4개 브라우저 병렬)
# ─────────────────────────────────────────────
async def _naver_web_worker(page, cats: list, short: str) -> int:
    cnt = 0
    for cat in cats:
        for suffix in ["010", "연락처"]:
            q = f"{short} {cat} {suffix}"
            for start in [1, 11, 21, 31, 41]:  # 5페이지
                url = f"https://search.naver.com/search.naver?query={urllib.parse.quote(q)}&where=web&start={start}"
                try:
                    await page.goto(url, timeout=18000)
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                    await asyncio.sleep(0.3)
                    text = await page.inner_text("body")
                    found = 0
                    for m in PHONE_RE.finditer(text):
                        if is_010(m.group(0)):
                            before = len(_records)
                            await add(m.group(0), cat, "네이버웹", cat)
                            if len(_records) > before:
                                cnt += 1
                                found += 1
                    # 새 번호 없으면 다음 페이지 안 봄
                    if found == 0 and start > 1:
                        break
                except Exception:
                    pass
                await asyncio.sleep(0.2)
    return cnt


async def scrape_naver_web(short: str) -> int:
    n = len(WEB_CATS)
    parts = [WEB_CATS[i*n//4:(i+1)*n//4] for i in range(4)]

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True, args=BROWSER_ARGS)
        ctxs  = [await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
                 for _ in range(4)]
        pages = [await ctx.new_page() for ctx in ctxs]
        for p in pages: await block_resources(p)
        r = await asyncio.gather(*[_naver_web_worker(pages[i], parts[i], short)
                                    for i in range(4)])
        await browser.close()
    return sum(r)


# ─────────────────────────────────────────────
# 5. DuckDuckGo (ddgs 라이브러리 + run_in_executor)
# ─────────────────────────────────────────────
def _ddg_sync(queries: list[str]) -> list[tuple[str, str]]:
    """동기 DDG 검색 — run_in_executor로 호출"""
    from ddgs import DDGS
    results = []
    ddgs = DDGS()
    for q in queries:
        try:
            for r in ddgs.text(q, region="kr-kr", max_results=20):
                text = r.get("title","") + " " + r.get("body","")
                results.append((q, text))
        except Exception:
            pass
    return results


async def scrape_ddg(short: str) -> int:
    queries = [f"{short} {cat} 010" for cat in WEB_CATS]
    loop = asyncio.get_event_loop()
    half = len(queries) // 2
    try:
        res_a, res_b = await asyncio.wait_for(asyncio.gather(
            loop.run_in_executor(None, _ddg_sync, queries[:half]),
            loop.run_in_executor(None, _ddg_sync, queries[half:]),
        ), timeout=120)  # 2분 max
    except asyncio.TimeoutError:
        print("  ⚠️ DDG 시간초과 (2분)")
        return 0
    cnt = 0
    for q, text in res_a + res_b:
        cat = q.split(" ")[-2] if len(q.split()) >= 2 else ""
        for m in PHONE_RE.finditer(text):
            if is_010(m.group(0)):
                before = len(_records)
                await add(m.group(0), cat, "DuckDuckGo", cat)
                if len(_records) > before: cnt += 1
    return cnt


# ─────────────────────────────────────────────
# 6. 네이버 place 검색 (where=place, 업체 카드에서 직접 추출)
# ─────────────────────────────────────────────
async def _naver_place_worker(page, cats: list, short: str) -> int:
    cnt = 0
    for cat in cats:
        q = f"{short} {cat}"
        url = f"https://search.naver.com/search.naver?query={urllib.parse.quote(q)}&where=place"
        try:
            await page.goto(url, timeout=18000)
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
            await asyncio.sleep(0.5)
            # 스크롤로 더 로드
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await asyncio.sleep(0.4)
            text = await page.inner_text("body")
            # 업체명도 같이 추출 시도
            items = await page.evaluate("""() => {
                const results = []
                // 네이버 place 결과 카드
                document.querySelectorAll('li[class*="place_item"], li[class*="PlaceItem"], div[class*="place_item"]').forEach(el => {
                    const name = el.querySelector('[class*="name"], [class*="title"], h3, h2')?.textContent?.trim() || ''
                    const phones = (el.innerText.match(/010[-\\s]?\\d{3,4}[-\\s]?\\d{4}/g) || [])
                    const tels = Array.from(el.querySelectorAll('a[href^="tel:"]')).map(a => a.href.replace('tel:',''))
                    phones.concat(tels).forEach(p => results.push({name, phone: p}))
                })
                return results
            }""")
            # items에서 추출
            for item in items:
                ph = item.get("phone", "")
                nm = item.get("name", "")
                if is_010(ph):
                    before = len(_records)
                    await add(ph, nm, "네이버place", cat)
                    if len(_records) > before: cnt += 1
            # fallback: body 전체 텍스트에서 정규식
            for m in PHONE_RE.finditer(text):
                if is_010(m.group(0)):
                    before = len(_records)
                    await add(m.group(0), cat, "네이버place", cat)
                    if len(_records) > before: cnt += 1
        except Exception:
            pass
        await asyncio.sleep(0.3)
    return cnt


async def scrape_naver_place(short: str) -> int:
    n = len(WEB_CATS)
    parts = [WEB_CATS[i*n//4:(i+1)*n//4] for i in range(4)]

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True, args=BROWSER_ARGS)
        ctxs  = [await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
                 for _ in range(4)]
        pages = [await ctx.new_page() for ctx in ctxs]
        for p in pages: await block_resources(p)
        r = await asyncio.gather(*[_naver_place_worker(pages[i], parts[i], short)
                                    for i in range(4)])
        await browser.close()
    return sum(r)


# ─────────────────────────────────────────────
# 네이버지도 검증 (옵션)
# ─────────────────────────────────────────────
def _addr_match(a1: str, a2: str) -> bool:
    """두 주소가 같은 위치인지 동/읍/면 레벨로 비교"""
    if not a1 or not a2:
        return True  # 주소 없으면 통과
    # 숫자·번지·층·호 제거 후 위치 키워드만 추출
    noise = re.compile(r'[\d\-]|층|호|관|번길|로|길')
    skip  = {'경기', '서울', '인천', '경남', '경북', '충남', '충북', '전남', '전북', '강원', '제주', '세종'}
    def tokens(a):
        return set(noise.sub(' ', a).split()) - skip - {''}
    return bool(tokens(a1) & tokens(a2))


async def _verify_chunk(page, chunk: list, region: str, progress: list):
    verified = 0
    for rec in chunk:
        name = rec.get("name", "")
        q = name if name else rec["phone"]
        try:
            await page.goto(
                f"https://m.search.naver.com/search.naver?query={urllib.parse.quote(q)}&where=m_local",
                timeout=18000
            )
            await page.wait_for_load_state("domcontentloaded", timeout=14000)
            await asyncio.sleep(0.4)
            html = await page.content()
            pid  = PLACE_ID_RE.search(html)
            if pid:
                nm = re.search(r'"placeName"\s*:\s*"([^"]{2,40})"', html)
                ad = re.search(r'"roadAddress"\s*:\s*"([^"]{5,80})"', html)
                rec["verified"]      = True
                rec["naver_url"]     = f"https://map.naver.com/p/entry/place/{pid.group(1)}"
                rec["naver_address"] = ad.group(1) if ad else ""
                if nm: rec["biz_name"] = nm.group(1)
                verified += 1
                progress[2][0] += 1
        except Exception:
            pass
        progress[0] += 1
        if progress[0] % 100 == 0:
            print(f"  [{progress[0]}/{progress[1]}] 확인됨: {progress[2][0]}개")
        await asyncio.sleep(0.4)
    return verified


async def verify_all(records: list, region: str):
    NUM = 5  # 병렬 브라우저 수
    chunks = [records[i::NUM] for i in range(NUM)]
    progress = [0, len(records), [0]]  # [done, total, [verified_count]]

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True, args=BROWSER_ARGS)
        pages = []
        for _ in range(NUM):
            ctx  = await browser.new_context(user_agent=MOBILE_UA, locale="ko-KR")
            page = await ctx.new_page()
            # 검증은 JS 렌더링 필요 → 리소스 차단 안 함
            pages.append(page)

        results = await asyncio.gather(
            *[_verify_chunk(pages[i], chunks[i], region, progress) for i in range(NUM)]
        )
        await browser.close()
    return sum(results)


# ─────────────────────────────────────────────
# 엑셀 저장
# ─────────────────────────────────────────────
def save_excel(records: list, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "010번호"

    hdrs = ["번호", "010번호", "업체명", "업종", "수집주소", "네이버지도주소", "검증", "네이버플레이스", "출처"]
    hfill = PatternFill("solid", fgColor="1a365d")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left",   vertical="center")

    ws.row_dimensions[1].height = 26
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = ctr

    green  = PatternFill("solid", fgColor="c6efce")
    yellow = PatternFill("solid", fgColor="ffeb9c")

    for idx, rec in enumerate(records, 1):
        row  = idx + 1
        fill = green if rec.get("verified") else yellow
        ws.row_dimensions[row].height = 20
        row_data = [
            idx, rec["phone"],
            rec.get("biz_name") or rec.get("name", ""),
            rec.get("category", ""),
            rec.get("collect_address", ""),
            rec.get("naver_address", ""),
            "✅" if rec.get("verified") else "⚠️",
            rec.get("naver_url", ""),
            rec.get("source", ""),
        ]
        lft_cols = {2,3,4,5,6,8,9}
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.alignment = lft if col in lft_cols else ctr
            if col == 8 and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    for col, w in enumerate([5,16,20,12,35,35,6,50,12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    Path(path).parent.mkdir(exist_ok=True)
    wb.save(path)


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
async def autosave_loop(region: str, stop_event: asyncio.Event):
    path = f"db/{region}_수집중.xlsx"
    while not stop_event.is_set():
        await asyncio.sleep(60)
        if _records:
            try:
                save_excel(list(_records.values()), path)
                print(f"  💾 자동저장: {len(_records)}개 → {path}")
            except Exception:
                pass


async def main():
    global _records
    _records = {}

    region = REGION
    short  = region.replace("경기도","").replace("시","").replace("군","").strip()

    areas = SUB_AREAS.get(region, [region])
    print("=" * 62)
    print(f"  소상공인 010번호 수집기 — 최종본")
    print(f"  지역: {region} ({len(areas)}개 서브지역)  검증: {'ON' if DO_VERIFY else 'OFF'}")
    print(f"  카카오: {len(areas)}지역 × {len(KAKAO_CATS)}카테고리 × 15페이지")
    print(f"  소스: 카카오맵(서브지역) + 당근 + 네이버웹(5p)")
    print("=" * 62)
    print()

    t0 = datetime.now()

    # ── 자동저장 시작 ───────────────────────────
    stop_event = asyncio.Event()
    if not ONLY:
        asyncio.create_task(autosave_loop(region, stop_event))

    # ── 소스 선택 실행 ──────────────────────────
    kakao_n = nmap_n = daangn_n = naver_n = 0

    if ONLY:
        print(f"🚀 단독 실행: {ONLY}")
        if ONLY == "kakao":    kakao_n  = await scrape_kakao(region)
        elif ONLY == "nmap":   nmap_n   = await scrape_naver_map(short)
        elif ONLY == "daangn": daangn_n = await scrape_daangn(short)
        elif ONLY == "web":    naver_n  = await scrape_naver_web(short)
        else: print(f"  알 수 없는 소스: {ONLY}")
    else:
        print("🚀 전체 수집 시작 (모든 소스 동시 실행)...")
        kakao_task   = asyncio.create_task(scrape_kakao(region))
        daangn_task  = asyncio.create_task(scrape_daangn(short))
        naver_task   = asyncio.create_task(scrape_naver_web(short))

        kakao_n, daangn_n, naver_n = await asyncio.gather(
            kakao_task, daangn_task, naver_task
        )

    stop_event.set()
    elapsed = (datetime.now() - t0).seconds
    records = list(_records.values())

    print(f"\n{'─'*62}")
    print(f"  수집 완료 ({elapsed//60}분 {elapsed%60}초)")
    print(f"  카카오맵  : +{kakao_n}개")
    print(f"  네이버맵  : +{nmap_n}개")
    print(f"  당근마켓  : +{daangn_n}개")
    print(f"  네이버웹  : +{naver_n}개")
    print(f"  총 (중복제거): {len(records)}개")
    print(f"{'─'*62}")

    # ── 검증 ────────────────────────────────────
    if DO_VERIFY and records:
        print(f"\n🗺️  네이버지도 검증 중... ({len(records)}개, 브라우저 5개 병렬)")
        print(f"  (스킵하려면 --no-verify 옵션 사용)")
        verified = await verify_all(records, region)
        print(f"  ✅ 확인됨: {verified}개 / 미확인: {len(records)-verified}개")

    # ── 저장 ────────────────────────────────────
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = f"db/{region}_010번호_최종_{ts}.xlsx"
    save_excel(records, out)

    print(f"\n{'='*62}")
    print(f"  완료!  파일: {out}")
    print(f"  총 010번호: {len(records)}개")
    print(f"{'='*62}")


if __name__ == "__main__":
    asyncio.run(main())
